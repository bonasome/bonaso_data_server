from django.test import TestCase

from rest_framework.test import APITestCase, APIClient
from rest_framework import status
from django.urls import reverse
from django.contrib.auth import get_user_model
from datetime import datetime
from projects.models import Project, Client, Task, Target
from respondents.models import Respondent, Interaction, InteractionSubcategory, HIVStatus, Pregnancy
from organizations.models import Organization
from indicators.models import Indicator, IndicatorSubcategory
from datetime import date
from io import BytesIO
from openpyxl import Workbook

User = get_user_model()

class UploadViewSetTest(APITestCase):
    def setUp(self):
        self.today = date.today().isoformat()
        
        self.admin = User.objects.create_user(username='testuser', password='testpass', role='admin')
        self.officer = User.objects.create_user(username='testuser2', password='testpass', role='meofficer')
        self.data_collector = User.objects.create_user(username='data_collector', password='testpass', role='data_collector')
        

        self.parent_org = Organization.objects.create(name='Test Org')
        self.child_org = Organization.objects.create(name='Test Org', parent_organization=self.parent_org)
        self.other_org = Organization.objects.create(name='Test Org')

        self.admin.organization = self.parent_org
        self.officer.organization = self.parent_org
        self.data_collector.organization = self.parent_org

        self.respondent_full = Respondent.objects.create(
            is_anonymous=False, 
            id_no= '1234567',
            first_name= 'Test',
            last_name= 'Testerson',
            dob= date(2000, 1, 1),
            ward= 'Here',
            village= 'ThePlace', 
            citizenship= 'Test',
            sex= Respondent.Sex.FEMALE,
            district= Respondent.District.CENTRAL,
        )
        
        self.project = Project.objects.create(
            name='Alpha Project',
            status=Project.Status.ACTIVE,
            start='2024-01-01',
            end='2024-12-31',
            description='Test project',
            created_by=self.admin,
        )
        self.project.organizations.set([self.parent_org, self.other_org])

        self.indicator = Indicator.objects.create(code='TEST1', name='Parent Indicator')
        self.child_indicator = Indicator.objects.create(code='TEST2', name='Child Indicator', prerequisite=self.indicator)
        
        self.project.indicators.set([self.indicator, self.child_indicator])

        self.task = Task.objects.create(project=self.project, organization=self.parent_org, indicator=self.indicator)
        self.prereq_task = Task.objects.create(project=self.project, organization=self.parent_org, indicator=self.child_indicator)

        self.child_task = Task.objects.create(project=self.project, organization=self.child_org, indicator=self.indicator)

        interaction = Interaction.objects.create(respondent=self.respondent_full, task=self.task, interaction_date='2024-05-01')




    def create_workbook(self, b1=None, b2=None, include_metadata=True, include_data=True):
        wb = Workbook()
        if include_metadata:
            ws = wb.active
            ws.title = 'Metadata'
            ws['B1'] = b1
            ws['B2'] = b2
        if include_data:
            wb.create_sheet(title='Data')
        return wb

    def test_no_file_uploaded(self):
        self.client.force_authenticate(user=self.officer)
        response = self.client.post('/api/record/interactions/upload/', {})
        self.assertEqual(response.status_code, status.HTTP_400_BAD_REQUEST)
        self.assertIn('No file was uploaded', str(response.data))

    def test_invalid_file_type(self):
        self.client.force_authenticate(user=self.officer)
        dummy_file = BytesIO(b"not excel")
        dummy_file.name = 'data.csv'
        response = self.client.post('/api/record/interactions/upload/', {'file': dummy_file}, format='multipart')
        self.assertEqual(response.status_code, status.HTTP_400_BAD_REQUEST)
        self.assertIn('must be an .xlsx', str(response.data))

    def test_missing_metadata_sheet(self):
        self.client.force_authenticate(user=self.officer)
        wb = Workbook()
        ws = wb.active
        ws.title = "NotMetadata"
        file_obj = BytesIO()
        wb.save(file_obj)
        file_obj.name = 'test.xlsx'
        file_obj.seek(0)

        response = self.client.post('/api/record/interactions/upload/', {'file': file_obj}, format='multipart')
        self.assertEqual(response.status_code, status.HTTP_400_BAD_REQUEST)
        self.assertIn("Unable to read 'Metadata'", str(response.data))

    def test_permission_denied_to_dc(self):
        self.client.force_authenticate(user=self.data_collector)
        wb = self.create_workbook(self.project.id, self.parent_org.id, include_data=True) 
        file_obj = BytesIO()
        wb.save(file_obj)
        file_obj.name = 'test.xlsx'
        file_obj.seek(0)

        response = self.client.post('/api/record/interactions/upload/', {'file': file_obj}, format='multipart')
        self.assertEqual(response.status_code, status.HTTP_403_FORBIDDEN)

    def test_permission_denied_other_org(self):
        self.client.force_authenticate(user=self.officer)
        wb = self.create_workbook(self.project.id, 99, include_data=True) 
        file_obj = BytesIO()
        wb.save(file_obj)
        file_obj.name = 'test.xlsx'
        file_obj.seek(0)

        response = self.client.post('/api/record/interactions/upload/', {'file': file_obj}, format='multipart')
        self.assertEqual(response.status_code, status.HTTP_403_FORBIDDEN)
    
    def test_missing_required_column(self):
        self.client.force_authenticate(user=self.officer)

        # Step 1: Create workbook with some but not all required headers
        wb = self.create_workbook(self.project.id, self.parent_org.id, include_data=True)  
        ws = wb['Data']
        # Missing 'First Name' on purpose
        headers = [
            "ID/Passport Number", "Last Name", "Age Range", "Date of Birth", "Sex", "Ward", "Village",
            "District", "Citizenship/Nationality", "HIV Status", "Date Positive", "Pregnant",
            "Email Address", "Phone Number", "Key Population Status", "Disability Status", "Date of Interaction",
            "Interaction Location"
        ]
        ws.append(headers)
        
        # Step 2: Save to in-memory file
        file_obj = BytesIO()
        wb.save(file_obj)
        file_obj.name = 'test.xlsx'
        file_obj.seek(0)

        # Step 3: Post to upload endpoint
        response = self.client.post('/api/record/interactions/upload/', {'file': file_obj}, format='multipart')

        # Step 4: Assert 400 response and expected error message
        self.assertEqual(response.status_code, status.HTTP_400_BAD_REQUEST)
        self.assertIn("Template is missing First Name column", str(response.data))





    def test_upload_respondents(self):
        self.client.force_authenticate(user=self.officer)
        wb = self.create_workbook(self.project.id, self.parent_org.id, include_data=True)  
        ws = wb['Data']
        # Missing 'First Name' on purpose
        headers = [
            "Is Anonymous","ID/Passport Number","First Name" , "Last Name", "Age Range", "Date of Birth", 
            "Sex", "Ward", "Village", "District", "Citizenship/Nationality", 
            "Email Address", "Phone Number", "Key Population Status",
            "Disability Status", "Special Respondent Attributes", "HIV Status", "Date Positive", "Pregnant",
            "Date of Interaction", "Interaction Location", "TEST1: Parent Indicator", 
            "TEST2: Child Indicator", "Comments"
        ]
        ws.append(headers)
        #ideal upload 
        row = [
            "FALSE", "T1", "Test", "Testerson", "", date(1990, 5, 1), "Male", "Wardplace", "Testington",
            "Central District", "Motswana", "test@website.com", "71234567", "Transgender, Intersex", 
            "Hearing Impaired, Visually Impaired", "Community Health Worker, Community Leader", 
            "Yes", date(2023,2,1), "", date(2024, 5, 1), "Mochudi", "Yes", "Yes", ""
        ]
        ws.append(row)
        #this is about as far of deviation as the function can handle, but should still upload
        row2 = [
            "FALSE", "T2", "Test", "Testerson", "", '6/25/2000', "male", "Wardplace", "Testington",
            "Central District", "Motswana", "", "", "transgender, Intersex", 
            "hearing   Impaired, Visually Impaired", " communityHealthworker,community   leader", "yes", "45447", "", '45447', "Mochudi", "Yes", "Yes", ""
        ]
        ws.append(row2)
        #test anon
        row3 = [
            "TRUE", "", "", "", "Under 18", "", "Female", "", "Testington",
            "Central District", "Motswana", "", "", "", 
            "", "", "", "", "Yes", date(2024, 5, 1), "Mochudi", "Yes", "Yes", ""
        ]
        ws.append(row3)

        file_obj = BytesIO()
        wb.save(file_obj)
        file_obj.name = 'template.xlsx'
        file_obj.seek(0)
        response = self.client.post('/api/record/interactions/upload/', {'file': file_obj}, format='multipart')
        print(response.json())
        self.assertEqual(response.status_code, status.HTTP_200_OK)

        response = self.client.get('/api/record/respondents/')
        self.assertEqual(response.status_code, status.HTTP_200_OK)
        self.assertEqual(len(response.data['results']), 4)

        respondent = Respondent.objects.get(id_no='T1')
        self.assertEqual(respondent.first_name, 'Test')
        self.assertEqual(respondent.kp_status.count(), 2)
        self.assertEqual(respondent.disability_status.count(), 2)
        self.assertEqual(respondent.special_attribute.count(), 2)
        hiv = HIVStatus.objects.get(respondent=respondent)
        self.assertEqual(hiv.date_positive, date(2023,2,1))

        #test should work
        respondent = Respondent.objects.get(id_no='T2')
        self.assertEqual(respondent.kp_status.count(), 2)
        self.assertEqual(respondent.disability_status.count(), 2)
        self.assertEqual(respondent.special_attribute.count(), 2)
        respondent = Respondent.objects.get(village='Testington', sex='F')
        preg = Pregnancy.objects.get(respondent=respondent)
        self.assertEqual(preg.term_began, date.today())

        
    def test_upload_respondents_child(self):
        self.client.force_authenticate(user=self.officer)
        wb = self.create_workbook(self.project.id, self.child_org.id, include_data=True)  
        ws = wb['Data']
        headers = [
            "Is Anonymous","ID/Passport Number","First Name" , "Last Name", "Age Range", "Date of Birth", 
            "Sex", "Ward", "Village", "District", "Citizenship/Nationality", 
            "Email Address", "Phone Number", "Key Population Status",
            "Disability Status", "HIV Status", "Date Positive", "Pregnant",
            "Date of Interaction", "Interaction Location", "TEST1: Parent Indicator", "TEST2: Child Indicator", "Comments"
        ]
        ws.append(headers)
        row = [
            "FALSE", "T1", "Test", "Testerson", "", date(1990, 5, 1), "Male", "Wardplace", "Testington",
            "Central District", "Motswana", "test@website.com", "71234567", "Transgender, Intersex", 
            "Hearing Impaired, Visually Impaired", "", "", "", date(2024, 5, 1), "Mochudi", "Yes", "Yes", ""
        ]
        ws.append(row)
        file_obj = BytesIO()
        wb.save(file_obj)
        file_obj.name = 'template.xlsx'
        file_obj.seek(0)
        response = self.client.post('/api/record/interactions/upload/', {'file': file_obj}, format='multipart')
        self.assertEqual(response.status_code, status.HTTP_200_OK)

        response = self.client.get('/api/record/respondents/')
        self.assertEqual(response.status_code, status.HTTP_200_OK)
        self.assertEqual(len(response.data['results']), 2)


    def test_missing_bad_values(self):
        #all of these should fail
        self.client.force_authenticate(user=self.officer)
        wb = self.create_workbook(self.project.id, self.parent_org.id, include_data=True)  
        ws = wb['Data']
        headers = [
            "Is Anonymous","ID/Passport Number","First Name" , "Last Name", "Age Range", "Date of Birth", 
            "Sex", "Ward", "Village", "District", "Citizenship/Nationality", 
            "Email Address", "Phone Number", "Key Population Status",
            "Disability Status", "HIV Status", "Date Positive", "Pregnant",
            "Date of Interaction", "Interaction Location", 
            "TEST1: Parent Indicator", "TEST2: Child Indicator", "Comments"
        ]
        ws.append(headers)
        #invalid options selected
        row = [
            "TRUE", "", "", "", "somewhere around 30, id say", "", "a guy", "Wardplace", "Testington",
            "you know that place, right?", "Motswana", "test@website.com", "71234567", "Transgender, Intersex", 
            "Hearing Impaired, Visually Impaired", "", "", "", date(2024, 5, 1), "Mochudi", "Yes", "Yes", ""
        ]
        ws.append(row)

        #missing values
        row2 = [
            "FALSE","T2", "Test", "", "", 'yesterday', "", "Wardplace", "Testington",
            "Central District", "Motswana", "test@website.com", "71234567", "Transgender, Intersex", 
            "Hearing Impaired, Visually Impaired", "", "", "", date(2024, 5, 1), "Mochudi", "Yes", "Yes", ""
        ]
        ws.append(row2)  

        #an non anon without a dob should also flag
        row3 = [
            "FALSE", "T2", "Test", "Testerson", "Under 18", "", "male", "Wardplace", "Testington",
            "Central District", "Motswana", "", "", "transgender, Intersex", 
            "hearing   Impaired; Visually Impaired", "", "", "", '45447', "Mochudi" "Yes", "Yes", ""
        ]
        ws.append(row3) 

        file_obj = BytesIO()
        wb.save(file_obj)
        file_obj.name = 'template.xlsx'
        file_obj.seek(0)

        #both of these should fail 
        response = self.client.post('/api/record/interactions/upload/', {'file': file_obj}, format='multipart')
        self.assertEqual(response.status_code, status.HTTP_200_OK)
        
        response = self.client.get('/api/record/respondents/')
        self.assertEqual(response.status_code, status.HTTP_200_OK)
        self.assertEqual(len(response.data['results']), 1)


    def test_bad_date(self):
        self.client.force_authenticate(user=self.officer)
        wb = self.create_workbook(self.project.id, self.parent_org.id, include_data=True)  
        ws = wb['Data']
        headers = [
           "Is Anonymous","ID/Passport Number","First Name" , "Last Name", "Age Range", "Date of Birth", 
            "Sex", "Ward", "Village", "District", "Citizenship/Nationality", 
            "Email Address", "Phone Number", "Key Population Status",
            "Disability Status", "HIV Status", "Date Positive", "Pregnant",
            "Date of Interaction", "Interaction Location", "TEST1: Parent Indicator", "TEST2: Child Indicator", "Comments"
        ]
        ws.append(headers)
        #future dates should block
        row = [
            "FALSE", "T1", "Test", "Testerson", "", date(2028, 5, 1), "Male", "Wardplace", "Testington",
            "Central District", "Motswana", "test@website.com", "71234567", "Transgender, Intersex", 
            "Hearing Impaired, Visually Impaired", "", "", "", date(2024, 5, 1), "Mochudi", "Yes", "Yes", ""
        ]
        ws.append(row)

        #as should unworkable dates
        row2 = [
            "FALSE","T2", "Test", "Testerson", "", 'yesterday', "Male", "Wardplace", "Testington",
            "Central District", "Motswana", "test@website.com", "71234567", "Transgender, Intersex", 
            "Hearing Impaired, Visually Impaired", "", "", "", date(2024, 5, 1), "Mochudi", "Yes", "Yes", ""
        ]
        ws.append(row2)  

        file_obj = BytesIO()
        wb.save(file_obj)
        file_obj.name = 'template.xlsx'
        file_obj.seek(0)

        #both of these should fail 
        response = self.client.post('/api/record/interactions/upload/', {'file': file_obj}, format='multipart')
        self.assertEqual(response.status_code, status.HTTP_200_OK)
        
        response = self.client.get('/api/record/respondents/')
        self.assertEqual(response.status_code, status.HTTP_200_OK)
        self.assertEqual(len(response.data['results']), 1)
    
    def test_bad_interaction_date(self):
        #respondent should save, but not interactions
        self.client.force_authenticate(user=self.officer)
        wb = self.create_workbook(self.project.id, self.parent_org.id, include_data=True)  
        ws = wb['Data']
        headers = [
            "Is Anonymous","ID/Passport Number","First Name" , "Last Name", "Age Range", "Date of Birth", 
            "Sex", "Ward", "Village", "District", "Citizenship/Nationality", 
            "Email Address", "Phone Number", "Key Population Status",
            "Disability Status", "HIV Status", "Date Positive", "Pregnant",
            "Date of Interaction", "Interaction Location", "TEST1: Parent Indicator", "TEST2: Child Indicator", "Comments"
        ]
        ws.append(headers)
        #this date is outside of the project range
        row = [
            "FALSE", "T1", "Test", "Testerson", "", date(2000, 5, 1), "Male", "Wardplace", "Testington",
            "Central District", "Motswana", "test@website.com", "71234567", "Transgender, Intersex", 
            "Hearing Impaired, Visually Impaired", "", "", "", date(2000, 5, 1), "Mochudi", "Yes", "Yes", ""
        ]
        ws.append(row)

        file_obj = BytesIO()
        wb.save(file_obj)
        file_obj.name = 'template.xlsx'
        file_obj.seek(0)

        #both of these should fail 
        response = self.client.post('/api/record/interactions/upload/', {'file': file_obj}, format='multipart')
        self.assertEqual(response.status_code, status.HTTP_200_OK)
        
        response = self.client.get('/api/record/respondents/')
        self.assertEqual(response.status_code, status.HTTP_200_OK)
        self.assertEqual(len(response.data['results']), 2)
        respondent = Respondent.objects.get(id_no='T1')
        interactions = Interaction.objects.filter(respondent=respondent).count()
        self.assertEqual(interactions, 0)
    
    def test_upload_same_respondent(self):
        #same respondents (same id number) should not be double recorded and should not throw a serious error
        #currently, we don't have override logic, server always wins
        self.client.force_authenticate(user=self.officer)
        wb = self.create_workbook(self.project.id, self.parent_org.id, include_data=True)  
        ws = wb['Data']
        # Missing 'First Name' on purpose
        headers = [
            "Is Anonymous","ID/Passport Number","First Name" , "Last Name", "Age Range", "Date of Birth", 
            "Sex", "Ward", "Village", "District", "Citizenship/Nationality", 
            "Email Address", "Phone Number", "Key Population Status",
            "Disability Status", "HIV Status", "Date Positive", "Pregnant",
            "Date of Interaction", "Interaction Location", "TEST1: Parent Indicator", "TEST2: Child Indicator", "Comments"
        ]
        ws.append(headers)
        #if respondent already exists/has interactions, these should be edited but not duplicated. 
        row = [
            "FALSE", "1234567", "Test", "Testerson", "", date(1990, 5, 1), "Male", "Wardplace", "Testington",
            "Central District", "Motswana", "test@website.com", "71234567", "Transgender, Intersex", 
            "Hearing Impaired, Visually Impaired", "", "", "", date(2024, 5, 1), "Mochudi", "Yes", "Yes", ""
        ]
        ws.append(row)

        file_obj = BytesIO()
        wb.save(file_obj)
        file_obj.name = 'template.xlsx'
        file_obj.seek(0)

        #both of these should fail 
        response = self.client.post('/api/record/interactions/upload/', {'file': file_obj}, format='multipart')
        self.assertEqual(response.status_code, status.HTTP_200_OK)
        
        response = self.client.get('/api/record/respondents/')
        self.assertEqual(response.status_code, status.HTTP_200_OK)
        self.assertEqual(len(response.data['results']), 1)

        interactions = Interaction.objects.filter(respondent=self.respondent_full).count()
        self.assertEqual(interactions, 2)
    

    def test_upload_same_respondent_update(self):
        #same respondents (same id number) should not be double recorded and should not throw a serious error
        #currently, we don't have override logic, server always wins
        numeric_ind = Indicator.objects.create(code ='NUM', name='Number', require_numeric=True)

        category1 = IndicatorSubcategory.objects.create(name='Cat 1', slug='cat1')
        category2 = IndicatorSubcategory.objects.create(name='Cat 2', slug='cat2')
        subcat_ind = Indicator.objects.create(code ='SC', name='Subcat')
      
        subcat_ind.subcategories.set([category1, category2])
        self.project.indicators.set([subcat_ind, numeric_ind])

        numeric_task = Task.objects.create(organization=self.parent_org, indicator=numeric_ind, project=self.project)
        subcat_task = Task.objects.create(organization=self.parent_org, indicator=subcat_ind, project=self.project)

        numeric_inter = Interaction.objects.create(respondent=self.respondent_full, interaction_date='2024-05-01', task=numeric_task, numeric_component=5)
        subcat_inter = Interaction.objects.create(respondent=self.respondent_full, interaction_date='2024-05-01', task=subcat_task)
        subcat_inter.subcategories.set([category1])

        self.client.force_authenticate(user=self.officer)
        wb = self.create_workbook(self.project.id, self.parent_org.id, include_data=True)  
        ws = wb['Data']
        # Missing 'First Name' on purpose
        headers = [
            "Is Anonymous","ID/Passport Number","First Name" , "Last Name", "Age Range", "Date of Birth", 
            "Sex", "Ward", "Village", "District", "Citizenship/Nationality", 
            "Email Address", "Phone Number", "Key Population Status",
            "Disability Status", "HIV Status", "Date Positive", "Pregnant",
            "Date of Interaction", "Interaction Location", "NUM: Number (Requires a Number)", "SC: Subcat", "Comments"
        ]
        ws.append(headers)
        #if respondent already exists/has interactions, these should be edited but not duplicated. 
        row = [
            "FALSE", "1234567", "Test", "Testerson", "", date(1990, 5, 1), "Male", "Wardplace", "Testington",
            "Central District", "Motswana", "test@website.com", "71234567", "Transgender, Intersex", 
            "Hearing Impaired, Visually Impaired", "", "", "", date(2024, 5, 1), "Mochudi", "10", "Cat 1, Cat 2", ""
        ]
        ws.append(row)

        file_obj = BytesIO()
        wb.save(file_obj)
        file_obj.name = 'template.xlsx'
        file_obj.seek(0)

        #both of these should fail 
        response = self.client.post('/api/record/interactions/upload/', {'file': file_obj}, format='multipart')
        self.assertEqual(response.status_code, status.HTTP_200_OK)
        
        response = self.client.get('/api/record/respondents/')
        self.assertEqual(response.status_code, status.HTTP_200_OK)
        self.assertEqual(len(response.data['results']), 1)

        interactions = Interaction.objects.filter(respondent=self.respondent_full).count()
        self.assertEqual(interactions, 3)
        int_detail = Interaction.objects.filter(respondent=self.respondent_full, task=numeric_task).first()
        self.assertEqual(int_detail.numeric_component, 10)
        int_detail = Interaction.objects.filter(respondent=self.respondent_full, task=subcat_task).first()
        self.assertEqual(int_detail.subcategories.count(), 2)
    
    def test_upload_interactions(self):
        #same respondents (same id number) should not be double recorded and should not throw a serious error
        #currently, we don't have override logic, server always wins
        self.client.force_authenticate(user=self.officer)
        wb = self.create_workbook(self.project.id, self.parent_org.id, include_data=True)  
        ws = wb['Data']
        # Missing 'First Name' on purpose
        headers = [
            "Is Anonymous","ID/Passport Number","First Name" , "Last Name", "Age Range", "Date of Birth", 
            "Sex", "Ward", "Village", "District", "Citizenship/Nationality", 
            "Email Address", "Phone Number", "Key Population Status",
            "Disability Status", "HIV Status", "Date Positive", "Pregnant",
            "Date of Interaction", "Interaction Location", "TEST1: Parent Indicator", "TEST2: Child Indicator", "Comments"
        ]
        ws.append(headers)
        row = [
            "FALSE", "T1", "Test", "Testerson", "", date(1990, 5, 1), "Male", "Wardplace", "Testington",
            "Central District", "Motswana", "test@website.com", "71234567", "Transgender, Intersex", 
            "Hearing Impaired, Visually Impaired", "", "", "", date(2024, 5, 1), "Mochudi", "Yes", "", ""
        ]
        ws.append(row)

        row2 = [
            "FALSE", "T2", "Test2", "Testerson II", "", date(1990, 5, 1), "Male", "Wardplace", "Testington",
            "Central District", "Motswana", "test@website.com", "71234567", "Transgender, Intersex", 
            "Hearing Impaired, Visually Impaired", "", "", "", date(2024, 5, 1), "Mochudi", "Yes", "Yes", ""
        ]
        ws.append(row2)

        file_obj = BytesIO()
        wb.save(file_obj)
        file_obj.name = 'template.xlsx'
        file_obj.seek(0)

        #both of these should fail 
        response = self.client.post('/api/record/interactions/upload/', {'file': file_obj}, format='multipart')
        self.assertEqual(response.status_code, status.HTTP_200_OK)
        
        response = self.client.get('/api/record/respondents/')
        self.assertEqual(response.status_code, status.HTTP_200_OK)
        self.assertEqual(len(response.data['results']), 3)
        respondent1 = Respondent.objects.get(id_no='T1')
        interactions1 = Interaction.objects.filter(respondent=respondent1).count()
        self.assertEqual(interactions1, 1)
        
        respondent2 = Respondent.objects.get(id_no='T2')
        interactions2 = Interaction.objects.filter(respondent=respondent2).count()
        self.assertEqual(interactions2, 2)
    

    def test_upload_interactions_no_prereq(self):
        #same respondents (same id number) should not be double recorded and should not throw a serious error
        #currently, we don't have override logic, server always wins
        self.client.force_authenticate(user=self.officer)
        wb = self.create_workbook(self.project.id, self.parent_org.id, include_data=True)  
        ws = wb['Data']
        #these interactions should fail since the parent indicator TEST1 has no interaction
        headers = [
            "Is Anonymous","ID/Passport Number","First Name" , "Last Name", "Age Range", "Date of Birth", 
            "Sex", "Ward", "Village", "District", "Citizenship/Nationality", 
            "Email Address", "Phone Number", "Key Population Status",
            "Disability Status", "HIV Status", "Date Positive", "Pregnant",
            "Date of Interaction", "Interaction Location", "TEST1: Parent Indicator", "TEST2: Child Indicator", "Comments"
        ]
        ws.append(headers)
        row = [
            "FALSE", "T1", "Test", "Testerson", "", date(1990, 5, 1), "Male", "Wardplace", "Testington",
            "Central District", "Motswana", "test@website.com", "71234567", "Transgender, Intersex", 
            "Hearing Impaired, Visually Impaired", "", "", "", date(2024, 5, 1), "Mochudi", "No", "Yes", ""
        ]
        ws.append(row)

        file_obj = BytesIO()
        wb.save(file_obj)
        file_obj.name = 'template.xlsx'
        file_obj.seek(0)

        #both of these should fail 
        response = self.client.post('/api/record/interactions/upload/', {'file': file_obj}, format='multipart')
        self.assertEqual(response.status_code, status.HTTP_200_OK)
        
        response = self.client.get('/api/record/respondents/')
        self.assertEqual(response.status_code, status.HTTP_200_OK)
        self.assertEqual(len(response.data['results']), 2)
        respondent = Respondent.objects.get(id_no='T1')
        interactions = Interaction.objects.filter(respondent=respondent).count()
        self.assertEqual(interactions, 0)
    
    def test_numeric(self):
        self.client.force_authenticate(user=self.officer)
        numeric_ind = Indicator.objects.create(code ='NUM', name='Number', require_numeric=True)

        self.project.indicators.set([numeric_ind])
        numeric_task = Task.objects.create(organization=self.parent_org, indicator=numeric_ind, project=self.project)

        wb = self.create_workbook(self.project.id, self.parent_org.id, include_data=True)  
        ws = wb['Data']
        #these interactions should fail since the parent indicator TEST1 has no interaction
        headers = [
            "Is Anonymous","ID/Passport Number","First Name" , "Last Name", "Age Range", "Date of Birth", 
            "Sex", "Ward", "Village", "District", "Citizenship/Nationality", 
            "Email Address", "Phone Number", "Key Population Status",
            "Disability Status", "HIV Status", "Date Positive", "Pregnant",
            "Date of Interaction", "Interaction Location", "NUM: Number (Requires a Number)", "Comments"
        ]
        ws.append(headers)
        #this is fine
        row = [
            "FALSE", "T1", "Test", "Testerson", "", date(1990, 5, 1), "Male", "Wardplace", "Testington",
            "Central District", "Motswana", "test@website.com", "71234567", "Transgender, Intersex", 
            "Hearing Impaired, Visually Impaired", "", "", "", date(2024, 5, 1), "Mochudi", "6", ""
        ]
        ws.append(row)

        file_obj = BytesIO()
        wb.save(file_obj)
        file_obj.name = 'template.xlsx'
        file_obj.seek(0)

        #both of these should fail 
        response = self.client.post('/api/record/interactions/upload/', {'file': file_obj}, format='multipart')
        self.assertEqual(response.status_code, status.HTTP_200_OK)
        
        response = self.client.get('/api/record/respondents/')
        self.assertEqual(response.status_code, status.HTTP_200_OK)
        self.assertEqual(len(response.data['results']), 2)
        respondent = Respondent.objects.get(id_no='T1')
        interactions = Interaction.objects.filter(respondent=respondent).count()
        int_detail = Interaction.objects.filter(respondent=respondent).first()
        self.assertEqual(interactions, 1)
        self.assertEqual(int_detail.numeric_component, 6)


    def test_subcats_valid(self):
        self.client.force_authenticate(user=self.officer)
        category1 = IndicatorSubcategory.objects.create(name='Cat 1', slug='cat1')
        category2 = IndicatorSubcategory.objects.create(name='Cat 2', slug='cat2')
        parent_ind = Indicator.objects.create(code ='SC', name='Subcat')
      
        parent_ind.subcategories.set([category1, category2])
        self.project.indicators.set([parent_ind])
        parent_task = Task.objects.create(organization=self.parent_org, indicator=parent_ind, project=self.project)

        wb = self.create_workbook(self.project.id, self.parent_org.id, include_data=True)  
        ws = wb['Data']
        #these interactions should fail since the parent indicator TEST1 has no interaction
        headers = [
            "Is Anonymous","ID/Passport Number","First Name" , "Last Name", "Age Range", "Date of Birth", 
            "Sex", "Ward", "Village", "District", "Citizenship/Nationality", 
            "Email Address", "Phone Number", "Key Population Status",
            "Disability Status", "HIV Status", "Date Positive", "Pregnant",
            "Date of Interaction", "Interaction Location", "SC: Subcat", "Comments"
        ]
        ws.append(headers)
        #this is fine
        row = [
            "FALSE", "T1", "Test", "Testerson", "", date(1990, 5, 1), "Male", "Wardplace", "Testington",
            "Central District", "Motswana", "test@website.com", "71234567", "Transgender, Intersex", 
            "Hearing Impaired, Visually Impaired", "", "", "", date(2024, 5, 1), "Mochudi", "Cat 1", ""
        ]
        ws.append(row)
        #should also work
        row2 = [
            "FALSE", "T2", "Test", "Testerson", "", date(1990, 5, 1), "Male", "Wardplace", "Testington",
            "Central District", "Motswana", "test@website.com", "71234567", "Transgender, Intersex", 
            "Hearing Impaired, Visually Impaired", "", "", "", date(2024, 5, 1), "Mochudi", "cat2,Cat1  ", ""
        ]
        ws.append(row2)

        file_obj = BytesIO()
        wb.save(file_obj)
        file_obj.name = 'template.xlsx'
        file_obj.seek(0)

        #both of these should fail 
        response = self.client.post('/api/record/interactions/upload/', {'file': file_obj}, format='multipart')
        self.assertEqual(response.status_code, status.HTTP_200_OK)
        
        response = self.client.get('/api/record/respondents/')
        self.assertEqual(response.status_code, status.HTTP_200_OK)
        self.assertEqual(len(response.data['results']), 3)
        respondent = Respondent.objects.get(id_no='T1')
        interactions = Interaction.objects.filter(respondent=respondent).count()
        int_detail = Interaction.objects.filter(respondent=respondent).first()
        self.assertEqual(interactions, 1)
        self.assertEqual(int_detail.subcategories.count(), 1)

        respondent = Respondent.objects.get(id_no='T2')
        interactions = Interaction.objects.filter(respondent=respondent).count()
        int_detail = Interaction.objects.filter(respondent=respondent).first()
        self.assertEqual(interactions, 1)
        self.assertEqual(int_detail.subcategories.count(), 2)

    def test_subcats_valid_numeric(self):
        self.client.force_authenticate(user=self.officer)
        category1 = IndicatorSubcategory.objects.create(name='Cat 1', slug='cat1')
        category2 = IndicatorSubcategory.objects.create(name='Cat 2', slug='cat2')
        parent_ind = Indicator.objects.create(code ='SC', name='Subcat', require_numeric=True)
      
        parent_ind.subcategories.set([category1, category2])
        self.project.indicators.set([parent_ind])
        parent_task = Task.objects.create(organization=self.parent_org, indicator=parent_ind, project=self.project)

        wb = self.create_workbook(self.project.id, self.parent_org.id, include_data=True)  
        ws = wb['Data']
        #these interactions should fail since the parent indicator TEST1 has no interaction
        headers = [
            "Is Anonymous","ID/Passport Number","First Name" , "Last Name", "Age Range", "Date of Birth", 
            "Sex", "Ward", "Village", "District", "Citizenship/Nationality", 
            "Email Address", "Phone Number", "Key Population Status",
            "Disability Status", "HIV Status", "Date Positive", "Pregnant",
            "Date of Interaction", "Interaction Location", "SC: Subcat (Requires a Number)", "Comments"
        ]
        ws.append(headers)
        #this is fine
        row = [
            "FALSE", "T1", "Test", "Testerson", "", date(1990, 5, 1), "Male", "Wardplace", "Testington",
            "Central District", "Motswana", "test@website.com", "71234567", "Transgender, Intersex", 
            "Hearing Impaired, Visually Impaired", "", "", "", date(2024, 5, 1), "Mochudi", "Cat 1: 12", ""
        ]
        ws.append(row)
        #should also work
        row2 = [
            "FALSE", "T2", "Test", "Testerson", "", date(1990, 5, 1), "Male", "Wardplace", "Testington",
            "Central District", "Motswana", "test@website.com", "71234567", "Transgender, Intersex", 
            "Hearing Impaired, Visually Impaired", "", "", "", date(2024, 5, 1), "Mochudi", "cat2:3,Cat1:5  ", ""
        ]
        ws.append(row2)

        file_obj = BytesIO()
        wb.save(file_obj)
        file_obj.name = 'template.xlsx'
        file_obj.seek(0)

        response = self.client.post('/api/record/interactions/upload/', {'file': file_obj}, format='multipart')
        self.assertEqual(response.status_code, status.HTTP_200_OK)
        
        response = self.client.get('/api/record/respondents/')
        self.assertEqual(response.status_code, status.HTTP_200_OK)
        self.assertEqual(len(response.data['results']), 3)
        respondent = Respondent.objects.get(id_no='T1')
        interactions = Interaction.objects.filter(respondent=respondent).count()
        int_detail = Interaction.objects.filter(respondent=respondent).first()
        self.assertEqual(interactions, 1)
        self.assertEqual(int_detail.subcategories.count(), 1)
        irsc1 = InteractionSubcategory.objects.get(interaction=int_detail, subcategory=category1)
        self.assertEqual(irsc1.numeric_component, 12)

        respondent = Respondent.objects.get(id_no='T2')
        interactions = Interaction.objects.filter(respondent=respondent).count()
        int_detail = Interaction.objects.filter(respondent=respondent).first()
        self.assertEqual(interactions, 1)
        self.assertEqual(int_detail.subcategories.count(), 2)
        irsc1 = InteractionSubcategory.objects.get(interaction=int_detail, subcategory=category1)
        self.assertEqual(irsc1.numeric_component, 5)
        irsc2 = InteractionSubcategory.objects.get(interaction=int_detail, subcategory=category2)
        self.assertEqual(irsc2.numeric_component, 3)

    def test_subcats_invalid_numeric(self):
        self.client.force_authenticate(user=self.officer)
        category1 = IndicatorSubcategory.objects.create(name='Cat 1', slug='cat1')
        category2 = IndicatorSubcategory.objects.create(name='Cat 2', slug='cat2')
        parent_ind = Indicator.objects.create(code ='SC', name='Subcat', require_numeric=True)
      
        parent_ind.subcategories.set([category1, category2])
        self.project.indicators.set([parent_ind])
        parent_task = Task.objects.create(organization=self.parent_org, indicator=parent_ind, project=self.project)

        wb = self.create_workbook(self.project.id, self.parent_org.id, include_data=True)  
        ws = wb['Data']
        #these interactions should fail since the parent indicator TEST1 has no interaction
        headers = [
            "Is Anonymous","ID/Passport Number","First Name" , "Last Name", "Age Range", "Date of Birth", 
            "Sex", "Ward", "Village", "District", "Citizenship/Nationality", 
            "Email Address", "Phone Number", "Key Population Status",
            "Disability Status", "HIV Status", "Date Positive", "Pregnant",
            "Date of Interaction", "Interaction Location", "SC: Subcat (Requires a Number)", "Comments"
        ]
        ws.append(headers)
        #this is fine
        row = [
            "FALSE", "T2", "Test", "Testerson", "", date(1990, 5, 1), "Male", "Wardplace", "Testington",
            "Central District", "Motswana", "test@website.com", "71234567", "Transgender, Intersex", 
            "Hearing Impaired, Visually Impaired", "", "", "", date(2024, 5, 1), "Mochudi", "Cat 1: p", ""
        ]
        ws.append(row)

        file_obj = BytesIO()
        wb.save(file_obj)
        file_obj.name = 'template.xlsx'
        file_obj.seek(0)

        response = self.client.post('/api/record/interactions/upload/', {'file': file_obj}, format='multipart')
        self.assertEqual(response.status_code, status.HTTP_200_OK)
        respondent = Respondent.objects.get(id_no='T2')
        interactions = Interaction.objects.filter(respondent=respondent).count()
        self.assertEqual(interactions, 0)

    def test_subcats_missing_numeric(self):
        self.client.force_authenticate(user=self.officer)
        category1 = IndicatorSubcategory.objects.create(name='Cat 1', slug='cat1')
        category2 = IndicatorSubcategory.objects.create(name='Cat 2', slug='cat2')
        parent_ind = Indicator.objects.create(code ='SC', name='Subcat', require_numeric=True)
      
        parent_ind.subcategories.set([category1, category2])
        self.project.indicators.set([parent_ind])
        parent_task = Task.objects.create(organization=self.parent_org, indicator=parent_ind, project=self.project)

        wb = self.create_workbook(self.project.id, self.parent_org.id, include_data=True)  
        ws = wb['Data']
        #these interactions should fail since the parent indicator TEST1 has no interaction
        headers = [
            "Is Anonymous","ID/Passport Number","First Name" , "Last Name", "Age Range", "Date of Birth", 
            "Sex", "Ward", "Village", "District", "Citizenship/Nationality", 
            "Email Address", "Phone Number", "Key Population Status",
            "Disability Status", "HIV Status", "Date Positive", "Pregnant",
            "Date of Interaction", "Interaction Location", "SC: Subcat (Requires a Number)", "Comments"
        ]
        ws.append(headers)
        #this is fine
        row = [
            "FALSE", "T2", "Test", "Testerson", "", date(1990, 5, 1), "Male", "Wardplace", "Testington",
            "Central District", "Motswana", "test@website.com", "71234567", "Transgender, Intersex", 
            "Hearing Impaired, Visually Impaired", "", "", "", date(2024, 5, 1), "Mochudi", "Cat 1", ""
        ]
        ws.append(row)

        file_obj = BytesIO()
        wb.save(file_obj)
        file_obj.name = 'template.xlsx'
        file_obj.seek(0)

        response = self.client.post('/api/record/interactions/upload/', {'file': file_obj}, format='multipart')
        self.assertEqual(response.status_code, status.HTTP_200_OK)
        respondent = Respondent.objects.get(id_no='T2')
        interactions = Interaction.objects.filter(respondent=respondent).count()
        self.assertEqual(interactions, 0)

    def test_subcats_invalid_missing(self):
        self.client.force_authenticate(user=self.officer)
        category1 = IndicatorSubcategory.objects.create(name='Cat 1', slug='cat1')
        category2 = IndicatorSubcategory.objects.create(name='Cat 2', slug='cat2')
        parent_ind = Indicator.objects.create(code ='SC', name='Subcat')
      
        parent_ind.subcategories.set([category1, category2])
        self.project.indicators.set([parent_ind])
        parent_task = Task.objects.create(organization=self.parent_org, indicator=parent_ind, project=self.project)

        wb = self.create_workbook(self.project.id, self.parent_org.id, include_data=True)  
        ws = wb['Data']
        #neither of these should log an interaction
        headers = [
            "Is Anonymous","ID/Passport Number","First Name" , "Last Name", "Age Range", "Date of Birth", 
            "Sex", "Ward", "Village", "District", "Citizenship/Nationality", 
            "Email Address", "Phone Number", "Key Population Status",
            "Disability Status", "HIV Status", "Date Positive", "Pregnant",
            "Date of Interaction", "Interaction Location", "SC: Subcat", "Comments"
        ]
        ws.append(headers)
        #blank will not log
        row = [
            "FALSE", "T1", "Test", "Testerson", "", date(1990, 5, 1), "Male", "Wardplace", "Testington",
            "Central District", "Motswana", "test@website.com", "71234567", "Transgender, Intersex", 
            "Hearing Impaired, Visually Impaired", "", "", "", date(2024, 5, 1), "Mochudi", "", ""
        ]
        ws.append(row)
        #forgot the comma
        row2 = [
            "FALSE", "T2", "Test", "Testerson", "", date(1990, 5, 1), "Male", "Wardplace", "Testington",
            "Central District", "Motswana", "test@website.com", "71234567", "Transgender, Intersex", 
            "Hearing Impaired, Visually Impaired", "", "", "", date(2024, 5, 1), "Mochudi", "cat1Cat1  ", ""
        ]
        ws.append(row2)

        file_obj = BytesIO()
        wb.save(file_obj)
        file_obj.name = 'template.xlsx'
        file_obj.seek(0)

        #both of these should fail 
        response = self.client.post('/api/record/interactions/upload/', {'file': file_obj}, format='multipart')
        self.assertEqual(response.status_code, status.HTTP_200_OK)
        
        response = self.client.get('/api/record/respondents/')
        self.assertEqual(response.status_code, status.HTTP_200_OK)
        self.assertEqual(len(response.data['results']), 3)
        respondent = Respondent.objects.get(id_no='T1')
        interactions = Interaction.objects.filter(respondent=respondent).count()
        self.assertEqual(interactions, 0)

        respondent = Respondent.objects.get(id_no='T2')
        interactions = Interaction.objects.filter(respondent=respondent).count()
        self.assertEqual(interactions, 0)

    

    def test_upload_mismatched_subcat(self):
        self.client.force_authenticate(user=self.officer)
        category1 = IndicatorSubcategory.objects.create(name='Cat 1', slug='cat1')
        category2 = IndicatorSubcategory.objects.create(name='Cat 2', slug='cat2')
        parent_ind = Indicator.objects.create(code ='PSC', name='Subcat Parent')
        child_ind = Indicator.objects.create(code ='CSC', name='Subcat Child', prerequisite=parent_ind)
      
        parent_ind.subcategories.set([category1, category2])
        child_ind.subcategories.set([category1, category2])
        self.project.indicators.set([parent_ind, child_ind])
        parent_task = Task.objects.create(organization=self.parent_org, indicator=parent_ind, project=self.project)
        child_task = Task.objects.create(organization=self.parent_org, indicator=child_ind, project=self.project)

        wb = self.create_workbook(self.project.id, self.parent_org.id, include_data=True)  
        ws = wb['Data']
        #these interactions should fail since the parent indicator TEST1 has no interaction
        headers = [
            "Is Anonymous","ID/Passport Number","First Name" , "Last Name", "Age Range", "Date of Birth", 
            "Sex", "Ward", "Village", "District", "Citizenship/Nationality", 
            "Email Address", "Phone Number", "Key Population Status",
            "Disability Status", "HIV Status", "Date Positive", "Pregnant",
            "Date of Interaction", "Interaction Location", "PSC: Subcat Parent", "CSC: Subcat Child", "Comments"
        ]
        ws.append(headers)
        row = [
            "FALSE", "T1", "Test", "Testerson", "", date(1990, 5, 1), "Male", "Wardplace", "Testington",
            "Central District", "Motswana", "test@website.com", "71234567", "Transgender, Intersex", 
            "Hearing Impaired, Visually Impaired", "", "", "", date(2024, 5, 1), "Mochudi", "Cat 1", "Cat 1, Cat 2", ""
        ]
        ws.append(row)

        file_obj = BytesIO()
        wb.save(file_obj)
        file_obj.name = 'template.xlsx'
        file_obj.seek(0)

        #both of these should fail 
        response = self.client.post('/api/record/interactions/upload/', {'file': file_obj}, format='multipart')
        self.assertEqual(response.status_code, status.HTTP_200_OK)
        
        response = self.client.get('/api/record/respondents/')
        self.assertEqual(response.status_code, status.HTTP_200_OK)
        self.assertEqual(len(response.data['results']), 2)
        respondent = Respondent.objects.get(id_no='T1')
        print(Interaction.objects.filter(respondent=respondent))
        interactions = Interaction.objects.filter(respondent=respondent).count()
        self.assertEqual(interactions, 1)
    
    def test_upload_matched_subcat(self):
        self.client.force_authenticate(user=self.officer)
        category1 = IndicatorSubcategory.objects.create(name='Cat 1', slug='cat1')
        category2 = IndicatorSubcategory.objects.create(name='Cat 2', slug='cat2')
        parent_ind = Indicator.objects.create(code ='PSC', name='Subcat Parent')
        child_ind = Indicator.objects.create(code ='CSC', name='Subcat Child', prerequisite=parent_ind)
      
        parent_ind.subcategories.set([category1, category2])
        child_ind.subcategories.set([category1, category2])
        self.project.indicators.set([parent_ind, child_ind])
        parent_task = Task.objects.create(organization=self.parent_org, indicator=parent_ind, project=self.project)
        child_task = Task.objects.create(organization=self.parent_org, indicator=child_ind, project=self.project)

        wb = self.create_workbook(self.project.id, self.parent_org.id, include_data=True)  
        ws = wb['Data']
        #these interactions should fail since the parent indicator TEST1 has no interaction
        headers = [
            "Is Anonymous","ID/Passport Number","First Name" , "Last Name", "Age Range", "Date of Birth", 
            "Sex", "Ward", "Village", "District", "Citizenship/Nationality", 
            "Email Address", "Phone Number", "Key Population Status",
            "Disability Status", "HIV Status", "Date Positive", "Pregnant",
            "Date of Interaction", "Interaction Location", "PSC: Subcat Parent", "CSC: Subcat Child", "Comments"
        ]
        ws.append(headers)
        row = [
            "FALSE", "T1", "Test", "Testerson", "", date(1990, 5, 1), "Male", "Wardplace", "Testington",
            "Central District", "Motswana", "test@website.com", "71234567", "Transgender, Intersex", 
            "Hearing Impaired, Visually Impaired", "", "", "", date(2024, 5, 1), "Mochudi", "Cat 1", "Cat 1", ""
        ]
        ws.append(row)

        file_obj = BytesIO()
        wb.save(file_obj)
        file_obj.name = 'template.xlsx'
        file_obj.seek(0)

        #both of these should fail 
        response = self.client.post('/api/record/interactions/upload/', {'file': file_obj}, format='multipart')
        self.assertEqual(response.status_code, status.HTTP_200_OK)
        
        response = self.client.get('/api/record/respondents/')
        self.assertEqual(response.status_code, status.HTTP_200_OK)
        self.assertEqual(len(response.data['results']), 2)
        respondent = Respondent.objects.get(id_no='T1')
        print(Interaction.objects.filter(respondent=respondent))
        interactions = Interaction.objects.filter(respondent=respondent).count()
        self.assertEqual(interactions, 2)