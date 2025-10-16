from django.test import TestCase

from rest_framework.test import APITestCase, APIClient
from rest_framework import status
from django.urls import reverse
from django.contrib.auth import get_user_model
from datetime import datetime
from events.models import Event
from projects.models import Project, Task, ProjectOrganization
from respondents.models import Respondent, Interaction, HIVStatus, Pregnancy, Response
from organizations.models import Organization
from indicators.models import Indicator, Assessment, LogicCondition, LogicGroup, Option
from datetime import date
from io import BytesIO
from openpyxl import Workbook

User = get_user_model()

class UploadViewSetTest(APITestCase):
    def setUp(self):
        self.today = date.today().isoformat()
        
        self.admin = User.objects.create_user(username='testuser', password='testpass', role='admin')
        self.officer = User.objects.create_user(username='testuser2', password='testpass', role='meofficer')
        self.data_collector = User.objects.create_user(username='data_collector', password='testpass', role='data_collector')
        

        self.parent_org = Organization.objects.create(name='Test Org')
        self.child_org = Organization.objects.create(name='Test Org')
        self.other_org = Organization.objects.create(name='Test Org')

        self.admin.organization = self.parent_org
        self.officer.organization = self.parent_org
        self.data_collector.organization = self.parent_org

        self.respondent_full = Respondent.objects.create(
            is_anonymous=False, 
            id_no= '1234567',
            first_name= 'Test',
            last_name= 'Testerson',
            dob= date(2000, 1, 1),
            ward= 'Here',
            village= 'ThePlace', 
            citizenship= 'Test',
            sex= Respondent.Sex.FEMALE,
            district= Respondent.District.CENTRAL,
        )
        
        self.project = Project.objects.create(
            name='Alpha Project',
            status=Project.Status.ACTIVE,
            start='2024-01-01',
            end='2024-12-31',
            description='Test project',
            created_by=self.admin,
        )
        self.project.organizations.set([self.parent_org, self.other_org, self.child_org])

        child_link = ProjectOrganization.objects.filter(organization=self.child_org).first()
        child_link.parent_organization = self.parent_org
        child_link.save()

        self.assessment = Assessment.objects.create(name='Ass')
        self.indicator = Indicator.objects.create(name='Screened for NCDs', type=Indicator.Type.MULTI, assessment=self.assessment, required=True)
        self.option1= Option.objects.create(name='BMI', indicator=self.indicator)
        self.option2= Option.objects.create(name='Blood Pressure', indicator=self.indicator)
        self.option3= Option.objects.create(name='Blood Glucose', indicator=self.indicator)

        self.indicator_log = Indicator.objects.create(name='Referred for NCDs', type=Indicator.Type.MULTI, assessment=self.assessment, match_options=self.indicator, allow_none=True)
        self.g2 = LogicGroup.objects.create(group_operator='AND', indicator=self.indicator_log)
        self.c21 = LogicCondition.objects.create(group=self.g2,source_type=LogicCondition.SourceType.ASS, source_indicator=self.indicator, condition_type=LogicCondition.ExtraChoices.ANY, operator=LogicCondition.Operator.EQUALS)
        self.indicator_sing = Indicator.objects.create(name='Type of Screening', required=True)
        self.option4= Option.objects.create(name='Type A', indicator=self.indicator_sing)
        self.option5= Option.objects.create(name='Type B', indicator=self.indicator_sing)
        self.indicator_num = Indicator.objects.create(name='Number of Sessions', required=True)


        self.task = Task.objects.create(project=self.project, organization=self.parent_org, assessment=self.assessment)
        self.child_task = Task.objects.create(project=self.project, organization=self.child_org, assessment=self.assessment)
        self.other_task = Task.objects.create(project=self.project, organization=self.other_org, assessment=self.assessment)

        self.interaction = Interaction.objects.create(interaction_date='2025-05-01', interaction_location='There', task=self.task, respondent=self.respondent_full)
        self.response = Response.objects.create(indicator=self.indicator, response_option=self.option1, interaction=self.interaction)
        
        self.headers = [
            "Is Anonymous","ID/Passport Number","First Name" , "Last Name", "Age Range", "Date of Birth", 
            "Sex", "Ward", "Village", "District", "Citizenship/Nationality", 
            "Email Address", "Phone Number", "Key Population Status",
            "Disability Status", "Special Respondent Attributes", "HIV Status", "Date Positive", "Pregnancy Began (Date)",
            "Pregnancy Ended (Date)", "Date of Interaction", "Interaction Location", 
            "Screened for NCDs: BMI (Select All That Apply)", "Screened for NCDs: Blood Pressure (Select All That Apply)", "Screened for NCDs: Blood Glucose (Select All That Apply)",
            "Referred for NCDs: BMI (Select All That Apply)", "Referred for NCDs: Blood Pressure (Select All That Apply)", "Referred for NCDs: Blood Glucose (Select All That Apply)",
            "Type of Screening (Select One)", "Number of Sessions (Enter a Number)", "Comments",
        ]

        self.resp_headers = ["FALSE", "T1", "Test", "Testerson", "", date(1990, 5, 1), "Male", "Wardplace", "Testington",
            "Central District", "BW", "test@website.com", "71234567", "Transgender, Intersex", 
            "Hearing Impaired, Visually Impaired", "Community Health Worker, Community Leader", 
            "Yes", date(2023,2,1), "", "", date(2024, 5, 1), "Mochudi"]
        
        self.resp_headers_2 = ["FALSE", "T2", "Test", "Testerson", "", date(1990, 5, 1), "Male", "Wardplace", "Testington",
            "Central District", "BW", "test@website.com", "71234567", "Transgender, Intersex", 
            "Hearing Impaired, Visually Impaired", "Community Health Worker, Community Leader", 
            "Yes", date(2023,2,1), "", "", date(2024, 5, 1), "Mochudi"]

    def create_workbook(self, tasks=None, org=None, include_metadata=True, include_data=True, include_event=False):
        '''
        Helper function that creates a sample workbook.
        '''
        wb = Workbook()
        if include_metadata:
            ws = wb.active
            ws.title = 'Metadata'
            ws['A2'] = org
            for i, task_id in enumerate(tasks):
                ws[f'B{i+2}'] = task_id
            ws["C1"] = "number of tasks"
            ws["C2"] = len(tasks)
        if include_data:
            wb.create_sheet(title='Data')
        return wb

    def test_no_file_uploaded(self):
        '''
        Test that an error is thrown if nothing is uploaded.
        '''
        self.client.force_authenticate(user=self.officer)
        response = self.client.post('/api/record/interactions/upload/', {})
        self.assertEqual(response.status_code, status.HTTP_400_BAD_REQUEST)
        self.assertIn('No file was uploaded', str(response.data))

    def test_invalid_file_type(self):
        '''
        Test that an error is thrown if the wrong file type is uploaded.
        '''
        self.client.force_authenticate(user=self.officer)
        dummy_file = BytesIO(b"not excel")
        dummy_file.name = 'data.csv'
        response = self.client.post('/api/record/interactions/upload/', {'file': dummy_file}, format='multipart')
        self.assertEqual(response.status_code, status.HTTP_400_BAD_REQUEST)
        self.assertIn('must be an .xlsx', str(response.data))

    def test_missing_metadata_sheet(self):
        '''
        Test that an error is thrown if the metadata sheet is missing.
        '''
        self.client.force_authenticate(user=self.officer)
        wb = Workbook()
        ws = wb.active
        ws.title = "NotMetadata"
        file_obj = BytesIO()
        wb.save(file_obj)
        file_obj.name = 'test.xlsx'
        file_obj.seek(0)

        response = self.client.post('/api/record/interactions/upload/', {'file': file_obj}, format='multipart')
        self.assertEqual(response.status_code, status.HTTP_400_BAD_REQUEST)
        self.assertIn("Unable to read 'Metadata'", str(response.data))

    def test_permission_denied_to_dc(self):
        '''
        Test that a dc cannot access templates.
        '''
        self.client.force_authenticate(user=self.data_collector)
        wb = self.create_workbook([self.task.id], self.parent_org.id, include_data=True) 
        file_obj = BytesIO()
        wb.save(file_obj)
        file_obj.name = 'test.xlsx'
        file_obj.seek(0)

        response = self.client.post('/api/record/interactions/upload/', {'file': file_obj}, format='multipart')
        self.assertEqual(response.status_code, status.HTTP_403_FORBIDDEN)

    def test_permission_denied_other_org(self):
        '''
        Test that you cannot access another orgs tempalte.
        '''
        self.client.force_authenticate(user=self.officer)
        wb = self.create_workbook([self.other_task.id], self.other_org, include_data=True) 
        file_obj = BytesIO()
        wb.save(file_obj)
        file_obj.name = 'test.xlsx'
        file_obj.seek(0)

        response = self.client.post('/api/record/interactions/upload/', {'file': file_obj}, format='multipart')
        self.assertEqual(response.status_code, status.HTTP_403_FORBIDDEN)
    
    def test_meta_mismatch(self):
        '''
        Test that you cannot access another orgs tempalte.
        '''
        self.client.force_authenticate(user=self.admin)
        wb = self.create_workbook([self.task.id], self.other_org, include_data=True) 
        file_obj = BytesIO()
        wb.save(file_obj)
        file_obj.name = 'test.xlsx'
        file_obj.seek(0)

        response = self.client.post('/api/record/interactions/upload/', {'file': file_obj}, format='multipart')
        self.assertEqual(response.status_code, status.HTTP_403_FORBIDDEN)
    
    def test_missing_required_column(self):
        '''
        Test that an error is thrown if a respondent information column is thrown.
        '''
        self.client.force_authenticate(user=self.officer)

        # Step 1: Create workbook with some but not all required headers
        wb = self.create_workbook([self.task.id], self.parent_org.id, include_data=True)  
        ws = wb['Data']
        # Missing 'First Name' on purpose
        headers = self.headers.remove('First Name')
        ws.append(headers)
        
        # Step 2: Save to in-memory file
        file_obj = BytesIO()
        wb.save(file_obj)
        file_obj.name = 'test.xlsx'
        file_obj.seek(0)

        # Step 3: Post to upload endpoint
        response = self.client.post('/api/record/interactions/upload/', {'file': file_obj}, format='multipart')

        # Step 4: Assert 400 response and expected error message
        self.assertEqual(response.status_code, status.HTTP_400_BAD_REQUEST)
        self.assertIn("Template is missing First Name column", str(response.data))


    def test_upload_success(self):
        '''
        Test that the file upload can successfuly create a respondent, including m2m fields, pregnancy,
        and HIV status.
        '''
        self.client.force_authenticate(user=self.officer)
        wb = self.create_workbook([self.task.id], self.parent_org.id, include_data=True)   
        ws = wb['Data']
        headers = self.headers
        ws.append(headers)
        #ideal upload 
        row = [
            "FALSE", "T1", "Test", "Testerson", "", date(1990, 5, 1), "Male", "Wardplace", "Testington",
            "Central District", "BW", "test@website.com", "71234567", "Transgender, Intersex", 
            "Hearing Impaired, Visually Impaired", "Community Health Worker, Community Leader", 
            "Yes", date(2023,2,1), "", "", date(2024, 5, 1), "Mochudi",

            "Yes", "Yes", "", 
            "Yes", "", "",
            "Type A",
            "7",
            "Random Comment",
        ]
        ws.append(row)
        #this is about as far of deviation as the function can handle, but should still upload
        row2 = [
            "FALSE", "T2", "Test", "Testerson", "", '6/25/2000', "male", "Wardplace", "Testington",
            "Central District", "BW", "", "", "transgender, Intersex", 
            "hearing   Impaired, Visually Impaired", " communityHealthworker,community   leader", "yes", "45447", "", "", 
            '45447', "Mochudi",

            "true", "blood pressure", "", 
            "yes", "", "",
            "type      a",
            "7  ",
            "Random Comment",
        ]
        ws.append(row2)
        #test anon
        row3 = [
            "TRUE", "", "", "", "20–24", "", "Female", "", "Testington",
            "Central District", "BW", "", "", "", 
            "", "", "", "", date(2023, 1, 1), date(2023, 9, 1), date(2024, 5, 1), "Mochudi",

            "Yes", "", "", 
            "Yes", "", "",
            "Type A",
            "7",
            "Random Comment",
        ]
        ws.append(row3)

        file_obj = BytesIO()
        wb.save(file_obj)
        file_obj.name = 'template.xlsx'
        file_obj.seek(0)
        response = self.client.post('/api/record/interactions/upload/', {'file': file_obj}, format='multipart')
        print(response.json())
        self.assertEqual(response.status_code, status.HTTP_200_OK)

        response = self.client.get('/api/record/respondents/')
        self.assertEqual(response.status_code, status.HTTP_200_OK)
        self.assertEqual(len(response.data['results']), 4) #these three plus the one that exists

        respondent = Respondent.objects.get(id_no='T1')
        self.assertEqual(respondent.first_name, 'Test')
        self.assertEqual(respondent.kp_status.count(), 2)
        self.assertEqual(respondent.disability_status.count(), 2)
        self.assertEqual(respondent.special_attribute.count(), 2)
        hiv = HIVStatus.objects.get(respondent=respondent)
        self.assertEqual(hiv.date_positive, date(2023,2,1))
        ir = Interaction.objects.filter(respondent=respondent)
        self.assertEqual(ir.interaction_date, date(2024, 5, 1))
        r1 = Response.objects.filter(interaction=ir, indicator=self.indicator)
        self.assertEqual(r1.count(), 2)
        r2 = Response.objects.filter(interaction=ir, indicator=self.indicator_log)
        self.assertEqual(r2.count(), 1)
        r3 = Response.objects.filter(interaction=ir, indicator=self.indicator_sing).first()
        self.assertEqual(r3.response_option, self.option4)
        r4 = Response.objects.filter(interaction=ir, indicator=self.indicator_num).first()
        self.assertEqual(r4.response_value, "7")
        self.assertEqual(ir.comments, "Random Comment")


        #test should work
        respondent = Respondent.objects.get(id_no='T2')
        self.assertEqual(respondent.kp_status.count(), 2)
        self.assertEqual(respondent.disability_status.count(), 2)
        self.assertEqual(respondent.special_attribute.count(), 2)
        ir = Interaction.objects.filter(respondent=respondent)
        self.assertEqual(ir.interaction_date, date(2024, 5, 1))
        r1 = Response.objects.filter(interaction=ir, indicator=self.indicator)
        self.assertEqual(r1.count(), 2)
        r2 = Response.objects.filter(interaction=ir, indicator=self.indicator_log)
        self.assertEqual(r2.count(), 1)
        r3 = Response.objects.filter(interaction=ir, indicator=self.indicator_sing).first()
        self.assertEqual(r3.response_option, self.option4)
        r4 = Response.objects.filter(interaction=ir, indicator=self.indicator_num).first()
        self.assertEqual(r3.response_value, "7")

        
        respondent = Respondent.objects.get(village='Testington', sex='F')
        preg = Pregnancy.objects.get(respondent=respondent)
        self.assertEqual(preg.term_began, date(2023, 1, 1))
        self.assertEqual(preg.term_ended, date(2023, 9, 1))
        self.assertEqual(preg.is_pregnant, False)

        
    def test_upload_respondents_child(self):
        '''
        Test that an M&E officer can upload a template for a child org.
        '''
        self.client.force_authenticate(user=self.officer)
        wb = self.create_workbook(self.project.id, self.child_org.id, include_data=True)  
        ws = wb['Data']
        headers = self.headers 
        ws.append(headers)
        row = self.resp_headers + ["Yes", "Yes", "", 
            "Yes", "", "",
            "Type A",
            "7",
            "Random Comment",
        ]
        ws.append(row)
        file_obj = BytesIO()
        wb.save(file_obj)
        file_obj.name = 'template.xlsx'
        file_obj.seek(0)
        response = self.client.post('/api/record/interactions/upload/', {'file': file_obj}, format='multipart')
        self.assertEqual(response.status_code, status.HTTP_200_OK)

        response = self.client.get('/api/record/respondents/')
        self.assertEqual(response.status_code, status.HTTP_200_OK)
        self.assertEqual(len(response.data['results']), 2)


    def test_missing_bad_values(self):
        '''
        Test that bad values are caught.
        '''
        self.client.force_authenticate(user=self.officer)
        wb = self.create_workbook(self.project.id, self.parent_org.id, include_data=True)  
        ws = wb['Data']
        headers = self.headers
        ws.append(headers)
        #invalid options selected
        row = [
            "TRUE", "", "", "", "somewhere around 30, id say", "", "a guy", "Wardplace", "Testington",
            "you know that place, right?", "BW", "test@website.com", "71234567", "Transgender, Intersex", 
            "Hearing Impaired, Visually Impaired", "", "", "", "", "", date(2024, 5, 1), "Mochudi", "Yes", "Yes", ""
        ]
        ws.append(row)

        #missing values
        row2 = [
            "FALSE","T2", "Test", "", "", 'yesterday', "", "Wardplace", "Testington",
            "Central District", "BW", "test@website.com", "71234567", "Transgender, Intersex", 
            "Hearing Impaired, Visually Impaired", "", "", "", "", "", date(2024, 5, 1), "Mochudi", "Yes", "Yes", ""
        ]
        ws.append(row2)  

        #an non anon without a dob should also flag
        row3 = [
            "FALSE", "T2", "Test", "Testerson", "Under 18", "", "male", "Wardplace", "Testington",
            "Central District", "BW", "", "", "transgender, Intersex", 
            "hearing   Impaired; Visually Impaired", "", "", "", "", "", '45447', "Mochudi" "Yes", "Yes", ""
        ]
        ws.append(row3) 

        file_obj = BytesIO()
        wb.save(file_obj)
        file_obj.name = 'template.xlsx'
        file_obj.seek(0)

        #both of these should fail 
        response = self.client.post('/api/record/interactions/upload/', {'file': file_obj}, format='multipart')
        print(response.json())
        self.assertEqual(response.status_code, status.HTTP_200_OK)
        
        response = self.client.get('/api/record/respondents/')
        self.assertEqual(response.status_code, status.HTTP_200_OK)
        self.assertEqual(len(response.data['results']), 1)


    def test_bad_date(self):
        '''
        Test date validations.
        '''
        self.client.force_authenticate(user=self.officer)
        wb = self.create_workbook(self.project.id, self.parent_org.id, include_data=True)  
        ws = wb['Data']
        headers = self.headers
        ws.append(headers)
        #future dates should block
        row = [
            "FALSE", "T1", "Test", "Testerson", "", date(2028, 5, 1), "Male", "Wardplace", "Testington",
            "Central District", "BW", "test@website.com", "71234567", "Transgender, Intersex", 
            "Hearing Impaired, Visually Impaired", "","", "", "", "", date(2024, 5, 1), "Mochudi", "Yes", "Yes", ""
        ]
        ws.append(row)

        #as should unworkable dates
        row2 = [
            "FALSE","T2", "Test", "Testerson", "", 'yesterday', "Male", "Wardplace", "Testington",
            "Central District", "BW", "test@website.com", "71234567", "Transgender, Intersex", 
            "Hearing Impaired, Visually Impaired", "","", "", "", "", date(2024, 5, 1), "Mochudi", "Yes", "Yes", ""
        ]
        ws.append(row2)  

        file_obj = BytesIO()
        wb.save(file_obj)
        file_obj.name = 'template.xlsx'
        file_obj.seek(0)

        #both of these should fail 
        response = self.client.post('/api/record/interactions/upload/', {'file': file_obj}, format='multipart')
        self.assertEqual(response.status_code, status.HTTP_200_OK)
        
        response = self.client.get('/api/record/respondents/')
        self.assertEqual(response.status_code, status.HTTP_200_OK)
        self.assertEqual(len(response.data['results']), 1)
    
    def test_bad_interaction_date(self):
        '''
        Test that an interaction date outside the project scope is caught. 
        Also, Respondent should save, but not interactions.
        '''
        self.client.force_authenticate(user=self.officer)
        wb = self.create_workbook(self.project.id, self.parent_org.id, include_data=True)  
        ws = wb['Data']
        headers = self.headers
        ws.append(headers)
        #this date is outside of the project range
        row = [
            "FALSE", "T1", "Test", "Testerson", "", date(2000, 5, 1), "Male", "Wardplace", "Testington",
            "Central District", "BW", "test@website.com", "71234567", "Transgender, Intersex", 
            "Hearing Impaired, Visually Impaired", "", "","", "", "", date(2000, 5, 1), "Mochudi",

            "Yes", "Yes", "", 
            "Yes", "", "",
            "Type A",
            "7",
            "Random Comment",
        ]
        ws.append(row)

        file_obj = BytesIO()
        wb.save(file_obj)
        file_obj.name = 'template.xlsx'
        file_obj.seek(0)

        #both of these should fail 
        response = self.client.post('/api/record/interactions/upload/', {'file': file_obj}, format='multipart')
        self.assertEqual(response.status_code, status.HTTP_200_OK)
        
        response = self.client.get('/api/record/respondents/')
        self.assertEqual(response.status_code, status.HTTP_200_OK)
        self.assertEqual(len(response.data['results']), 2)
        respondent = Respondent.objects.get(id_no='T1')
        interactions = Interaction.objects.filter(respondent=respondent).count()
        self.assertEqual(interactions, 0)
    
    def test_bad_logic(self):
        '''
        Test that the file upload can successfuly create a respondent, including m2m fields, pregnancy,
        and HIV status.
        '''
        self.client.force_authenticate(user=self.officer)
        wb = self.create_workbook([self.task.id], self.parent_org.id, include_data=True)   
        ws = wb['Data']
        headers = self.headers
        ws.append(headers)
        #ideal upload 
        row = [
            "FALSE", "T1", "Test", "Testerson", "", date(1990, 5, 1), "Male", "Wardplace", "Testington",
            "Central District", "BW", "test@website.com", "71234567", "Transgender, Intersex", 
            "Hearing Impaired, Visually Impaired", "Community Health Worker, Community Leader", 
            "Yes", date(2023,2,1), "", "", date(2024, 5, 1), "Mochudi",

            "Yes", "", "", 
            "Yes", "Yes", "",
            "Type A",
            "7",
            "Random Comment",
        ]
        ws.append(row)

        row2 = [
            "FALSE", "T1", "Test", "Testerson", "", date(1990, 5, 1), "Male", "Wardplace", "Testington",
            "Central District", "BW", "test@website.com", "71234567", "Transgender, Intersex", 
            "Hearing Impaired, Visually Impaired", "Community Health Worker, Community Leader", 
            "Yes", date(2023,2,1), "", "", date(2024, 5, 1), "Mochudi",

            "", "", "", 
            "Yes", "", "",
            "Type A",
            "7",
            "Random Comment",
        ]
        ws.append(row2)
        file_obj = BytesIO()
        wb.save(file_obj)
        file_obj.name = 'template.xlsx'
        file_obj.seek(0)

        #both of these should fail 
        response = self.client.post('/api/record/interactions/upload/', {'file': file_obj}, format='multipart')
        self.assertEqual(response.status_code, status.HTTP_200_OK)

        ir = Interaction.objects.all()
        self.assertEqual(ir.count(), 1)
        
    def test_upload_same_respondent(self):
        '''
        Test that the same respondent is caught and not recreated. The frontend will give a display for this.
        '''
        self.client.force_authenticate(user=self.officer)
        wb = self.create_workbook(self.project.id, self.parent_org.id, include_data=True)  
        ws = wb['Data']
        # Missing 'First Name' on purpose
        headers = self.headers
        ws.append(headers)
        #if respondent already exists/has interactions, these should be edited but not duplicated. 
        row = [
            "FALSE", "1234567", "Test", "Testerson", "", date(1990, 5, 1), "Male", "Wardplace", "Testington",
            "Central District", "BW", "test@website.com", "71234567", "Transgender, Intersex", 
            "Hearing Impaired, Visually Impaired", "","", "", "", "", date(2024, 5, 1), "Mochudi"
            
            "Yes", "Yes", "", 
            "Yes", "", "",
            "Type A",
            "7",
            "Random Comment",
        ]
        ws.append(row)

        file_obj = BytesIO()
        wb.save(file_obj)
        file_obj.name = 'template.xlsx'
        file_obj.seek(0)

        #both of these should fail 
        response = self.client.post('/api/record/interactions/upload/', {'file': file_obj}, format='multipart')
        print(response.json())
        self.assertEqual(response.status_code, status.HTTP_200_OK)
        self.assertEqual(len(response.data['conflicts']), 1)
        response = self.client.get('/api/record/respondents/')
        self.assertEqual(response.status_code, status.HTTP_200_OK)
        self.assertEqual(len(response.data['results']), 1)

        interactions = Interaction.objects.filter(respondent=self.respondent_full).count()
        self.assertEqual(interactions, 1)
        responses = Response.objects.all()
        self.assertEqual(responses.count(), 5)
    
    def test_citizenship(self):
        '''
        Test that the same respondent is caught and not recreated. The frontend will give a display for this.
        '''
        self.client.force_authenticate(user=self.officer)
        wb = self.create_workbook(self.project.id, self.parent_org.id, include_data=True)  
        ws = wb['Data']
        # Missing 'First Name' on purpose
        headers = self.headers + ["TEST1: Parent Indicator", "TEST2: Child Indicator", "Comments"]
        ws.append(headers)
        #name should work 
        row = [
            "FALSE", "000010000", "Test", "Testerson", "", date(1990, 5, 1), "Male", "Wardplace", "Testington",
            "Central District", "botswana", "test@website.com", "71234567", "Transgender, Intersex", 
            "Hearing Impaired, Visually Impaired", "","", "", "", "", date(2024, 5, 1), "Mochudi", 
            "Yes", "Yes", "", 
            "Yes", "", "",
            "Type A",
            "7",
            "Random Comment",
        ]
        ws.append(row)


        row2 = [
            "FALSE", "100010000", "Test", "Testerson", "", date(1990, 5, 1), "Male", "Wardplace", "Testington",
            "Central District", "USA", "test@website.com", "71234567", "Transgender, Intersex", 
            "Hearing Impaired, Visually Impaired", "","", "", "", "", date(2024, 5, 1), "Mochudi", 
            "Yes", "Yes", "", 
            "Yes", "", "",
            "Type A",
            "7",
            "Random Comment",
        ]
        ws.append(row2)

        row3 = [
            "FALSE", "200010000", "Test", "Testerson", "", date(1990, 5, 1), "Male", "Wardplace", "Testington",
            "Central District", "pxQw2", "test@website.com", "71234567", "Transgender, Intersex", 
            "Hearing Impaired, Visually Impaired", "","", "", "", "", date(2024, 5, 1), "Mochudi", 
            "Yes", "Yes", "", 
            "Yes", "", "",
            "Type A",
            "7",
            "Random Comment",
        ]
        ws.append(row3)

        file_obj = BytesIO()
        wb.save(file_obj)
        file_obj.name = 'template.xlsx'
        file_obj.seek(0)

        #both of these should fail 
        response = self.client.post('/api/record/interactions/upload/', {'file': file_obj}, format='multipart')
        print(response.json())
        self.assertEqual(response.status_code, status.HTTP_200_OK)
        r1 = Respondent.objects.filter(id_no='000010000').first()
        self.assertEqual(r1.citizenship, 'BW')
        r2 = Respondent.objects.filter(id_no='100010000').first()
        self.assertEqual(r2.citizenship, 'US')
