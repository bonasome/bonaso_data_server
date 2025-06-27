from django.test import TestCase

from rest_framework.test import APITestCase, APIClient
from rest_framework import status
from django.urls import reverse
from django.contrib.auth import get_user_model
from datetime import datetime
from projects.models import Project, Client, Task, Target
from respondents.models import Respondent, Interaction, Pregnancy, HIVStatus
from organizations.models import Organization
from indicators.models import Indicator, IndicatorSubcategory
from datetime import date
from io import BytesIO
from openpyxl import Workbook
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter, column_index_from_string

User = get_user_model()

class TemplateActionTest(APITestCase):
    def setUp(self):
        self.today = date.today().isoformat()
        
        self.admin = User.objects.create_user(username='testuser', password='testpass', role='admin')
        self.officer = User.objects.create_user(username='testuser2', password='testpass', role='meofficer')
        self.data_collector = User.objects.create_user(username='data_collector', password='testpass', role='data_collector')
        self.other = User.objects.create_user(username='loser', password='testpass', role='meofficer')

        self.parent_org = Organization.objects.create(name='Test Org')
        self.child_org = Organization.objects.create(name='Test Org', parent_organization=self.parent_org)
        self.other_org = Organization.objects.create(name='Test Org')

        self.admin.organization = self.parent_org
        self.officer.organization = self.parent_org
        self.data_collector.organization = self.parent_org
        self.other.organization = self.other_org
        
        self.project = Project.objects.create(
            name='Alpha Project',
            status=Project.Status.ACTIVE,
            start='2024-01-01',
            end='2024-12-31',
            description='Test project',
            created_by=self.admin,
        )
        self.project.organizations.set([self.parent_org, self.other_org])

        self.indicator = Indicator.objects.create(code='TEST1', name='Parent Indicator')
        self.child_indicator = Indicator.objects.create(code='TEST2', name='Child Indicator', prerequisite=self.indicator)
        self.numeric_indicator = Indicator.objects.create(code='NUM', name='Number', require_numeric=True)
        self.subcat_indicator = Indicator.objects.create(code='SC', name='Subcat')
        category1 = IndicatorSubcategory.objects.create(name='Cat 1')
        category2 = IndicatorSubcategory.objects.create(name='Cat 2')
        self.subcat_indicator.subcategories.set([category1, category2])

        self.project.indicators.set([self.indicator, self.child_indicator, self.numeric_indicator, self.subcat_indicator])
        

        self.task = Task.objects.create(project=self.project, organization=self.parent_org, indicator=self.indicator)
        self.prereq_task = Task.objects.create(project=self.project, organization=self.parent_org, indicator=self.child_indicator)
        self.numeric_task = Task.objects.create(project=self.project, organization=self.parent_org, indicator=self.numeric_indicator)
        self.subcat_task = Task.objects.create(project=self.project, organization=self.parent_org, indicator=self.subcat_indicator)
        
        self.child_task = Task.objects.create(project=self.project, organization=self.child_org, indicator=self.indicator)
    def get_column_index_by_header(self, sheet, header_name):
        """Returns 1-based column index for a given header name in the first row."""
        for cell in sheet[1]:
            if cell.value == header_name:
                return cell.column  # returns 1-based index like 1 for A, 2 for B, etc.
        raise ValueError(f"Header '{header_name}' not found in first row.")

    def test_correct_template(self):
        self.client.force_authenticate(user=self.admin)
        response = self.client.get('/api/record/interactions/template/', {
            'project': self.project.id,
            'organization': self.parent_org.id
        })
        self.assertEqual(response.status_code, 200)
        self.assertEqual(response['Content-Type'], 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')

        # Confirm workbook structure (respondent fields + tasks)
        wb = load_workbook(filename=BytesIO(response.content))
        self.assertIn('Data', wb.sheetnames)
        self.assertIn('DropdownOptions', wb.sheetnames)
        self.assertIn('Metadata', wb.sheetnames)

        #confirm metadata sheet is present
        metadata = wb['Metadata']
        self.assertEqual(metadata['B1'].value, str(self.project.id))
        self.assertEqual(metadata['B2'].value, str(self.parent_org.id))
        data=wb['Data']
        actual_headers = [cell.value for cell in data[1] if cell.value]

        expected_headers = [
            'Is Anonymous',

            'ID/Passport Number',
            'First Name',
            'Last Name',
            'Sex',
            'Date of Birth',
            'District',
            'Village',
            'Ward',
            'Citizenship/Nationality',
            'Comments',
            'Age Range',
            'Key Population Status',
            'Disability Status',
            'Date of Interaction',
            f'{self.indicator.code}: {self.indicator.name}',
            f'{self.child_indicator.code}: {self.child_indicator.name}',
            f'{self.numeric_indicator.code}: {self.numeric_indicator.name} (Requires a Number)',
            f'{self.subcat_indicator.code}: {self.subcat_indicator.name}',

        ]
        for header in expected_headers:
            self.assertIn(header, actual_headers, msg=f"Missing expected header: {header}")
        
        #confirm data validation sheet is correctly set up
        options = wb['DropdownOptions']
        actual_headers = [cell.value for cell in options[1] if cell.value]

        for header in expected_headers:
            #6 dropdown options for respondent fields
            #each non-numeric task should also have a list, either of subcats or a straight yes/no
            self.assertEqual(len(actual_headers), 9)
        
        name=f'{self.subcat_indicator.code}: {self.subcat_indicator.name}'
        subcat_col = self.get_column_index_by_header(data, name)
        col_letter = get_column_letter(subcat_col)
        self.assertEqual(options[f'{col_letter}1'].value, 'Cat 1')
        self.assertEqual(options[f'{col_letter}2'].value, 'Cat 2')
    
    def test_me_mgr_access(self):
        #managers/me officers can access templates for their organization
        self.client.force_authenticate(user=self.officer)
        response = self.client.get('/api/record/interactions/template/', {
            'project': self.project.id,
            'organization': self.parent_org.id
        })
        self.assertEqual(response.status_code, 200)
    
    def test_me_mgr_access_child(self):
        #and their child orgs
        self.client.force_authenticate(user=self.officer)
        response = self.client.get('/api/record/interactions/template/', {
            'project': self.project.id,
            'organization': self.child_org.id
        })
        self.assertEqual(response.status_code, 200)
    
    def test_access_wrong_org(self):
        #but not other orgs
        self.client.force_authenticate(user=self.officer)
        response = self.client.get('/api/record/interactions/template/', {
            'project': self.project.id,
            'organization': self.other_org.id
        })
        self.assertEqual(response.status_code, 403)
    
    def test_no_task(self):
        #if a project has no tasks, there's no need for a template, so return a 400
        self.client.force_authenticate(user=self.other)
        response = self.client.get('/api/record/interactions/template/', {
            'project': self.project.id,
            'organization': self.other_org.id
        })
        self.assertEqual(response.status_code, 400)
    
    def test_no_perm(self):
        #data collectors are not allowed to access templates
        self.client.force_authenticate(user=self.data_collector)
        response = self.client.get('/api/record/interactions/template/', {
            'project': self.project.id,
            'organization': self.parent_org.id
        })
        self.assertEqual(response.status_code, 403)

        