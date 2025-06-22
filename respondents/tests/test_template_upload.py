from io import BytesIO
from openpyxl import Workbook
from django.urls import reverse
from rest_framework.test import APITestCase, APIClient
from rest_framework import status
from django.contrib.auth import get_user_model
from organizations.models import Organization
from projects.models import Project, Task, Client
from indicators.models import Indicator, IndicatorSubcategory
from datetime import date
User = get_user_model()

class TemplateUploadTests(APITestCase):
    def setUp(self):
        self.client = APIClient()
        self.org = Organization.objects.create(name='Parent Org')

        self.client_obj = Client.objects.create(name='Test Client')
        self.project = Project.objects.create(
            name='Beta Project',
            client=self.client_obj,
            status=Project.Status.PLANNED,
            start='2024-02-01',
            end='2024-10-31',
            description='Second project',
        )
        self.project.organizations.set([self.org])

        self.indicator = Indicator.objects.create(code='1', name='Test Ind')
        self.indicator_subcat = Indicator.objects.create(code='2', name='Has Subcats')
        subcats = ['CCC', 'HIV', 'NCD']
        subcat_obj = []
        for c in subcats:
            obj = IndicatorSubcategory.objects.create(name=c)
            subcat_obj.append(obj)
        self.indicator.subcategories.set(subcat_obj)
        self.indicator_prereq = Indicator.objects.create(code='3', name='Requires 1', prerequisite=self.indicator)

        self.project.indicators.set([self.indicator, self.indicator_subcat, self.indicator_prereq])

        self.task = Task.objects.create(project=self.project, organization=self.org, indicator=self.indicator)
        self.task_subcat = Task.objects.create(project=self.project, organization=self.org, indicator=self.indicator_subcat)
        self.task_prereq = Task.objects.create(project=self.project, organization=self.org, indicator=self.indicator_prereq)

        self.user = User.objects.create_user(username='manager', password='pass', role='manager', organization=self.org)
        self.no_perm_user = User.objects.create_user(username='dc', password='pass', role='data_collector', organization=self.org)
        
        self.url = reverse('interaction-post-template') 

    def create_workbook(self, b1=None, b2=None, include_metadata=True, include_data=True):
        self.client.force_authenticate(user=self.user)
        wb = Workbook()
        if include_metadata:
            ws = wb.active
            ws.title = 'Metadata'
            ws['B1'] = b1
            ws['B2'] = b2
        if include_data:
            wb.create_sheet(title='Data')
        return wb

    def test_no_file_uploaded(self):
        self.client.force_authenticate(user=self.user)
        response = self.client.post(self.url, {})
        self.assertEqual(response.status_code, status.HTTP_400_BAD_REQUEST)
        self.assertIn('No file was uploaded', str(response.data))

    def test_invalid_file_type(self):
        self.client.force_authenticate(user=self.user)
        dummy_file = BytesIO(b"not excel")
        dummy_file.name = 'data.csv'
        response = self.client.post(self.url, {'file': dummy_file}, format='multipart')
        self.assertEqual(response.status_code, status.HTTP_400_BAD_REQUEST)
        self.assertIn('must be an .xlsx', str(response.data))

    def test_missing_metadata_sheet(self):
        self.client.force_authenticate(user=self.user)
        wb = Workbook()
        ws = wb.active
        ws.title = "NotMetadata"
        file_obj = BytesIO()
        wb.save(file_obj)
        file_obj.name = 'test.xlsx'
        file_obj.seek(0)

        response = self.client.post(self.url, {'file': file_obj}, format='multipart')
        self.assertEqual(response.status_code, status.HTTP_400_BAD_REQUEST)
        self.assertIn("Unable to read 'Metadata'", str(response.data))

    def test_non_numeric_ids(self):
        self.client.force_authenticate(user=self.user)
        wb = self.create_workbook('abc', 'xyz')
        file_obj = BytesIO()
        wb.save(file_obj)
        file_obj.name = 'test.xlsx'
        file_obj.seek(0)

        response = self.client.post(self.url, {'file': file_obj}, format='multipart')
        self.assertEqual(response.status_code, status.HTTP_400_BAD_REQUEST)
        self.assertIn('must be numeric', str(response.data))

    def test_permission_denied_to_other_org(self):
        self.client.force_authenticate(user=self.user)
        wb = self.create_workbook(1, 999)  
        file_obj = BytesIO()
        wb.save(file_obj)
        file_obj.name = 'test.xlsx'
        file_obj.seek(0)

        response = self.client.post(self.url, {'file': file_obj}, format='multipart')
        self.assertEqual(response.status_code, status.HTTP_403_FORBIDDEN)
    
    def test_missing_required_column(self):
        self.client.force_authenticate(user=self.user)

        # Step 1: Create workbook with some but not all required headers
        wb = self.create_workbook(self.project.id, self.org.id, include_data=True)  
        ws = wb['Data']
        # Missing 'First Name' on purpose
        headers = [
            "ID/Passport Number", "Last Name", "Age Range", "Date of Birth", "Sex", "Ward", "Village",
            "District", "Citizenship/Nationality", "Email Address", "Phone Number", "Key Population Status",
            "Disability Status", "Date of Interaction"
        ]
        ws.append(headers)
        
        # Step 2: Save to in-memory file
        file_obj = BytesIO()
        wb.save(file_obj)
        file_obj.name = 'test.xlsx'
        file_obj.seek(0)

        # Step 3: Post to upload endpoint
        response = self.client.post(self.url, {'file': file_obj}, format='multipart')

        # Step 4: Assert 400 response and expected error message
        self.assertEqual(response.status_code, status.HTTP_400_BAD_REQUEST)
        self.assertIn("Template is missing First Name column", str(response.data))
    
    def test_invalid_respondent_row_validation(self):
        self.client.force_authenticate(user=self.user)

        # Create an Excel file with a single invalid row
        wb = self.create_workbook(self.project.id, self.org.id, include_data=True)  
        ws = wb['Data']

        headers = [
            "ID/Passport Number", "First Name", "Last Name", "Age Range", "Date of Birth", "Sex",
            "Ward", "Village", "District", "Citizenship", "Email Address", "Phone Number",
            "Key Population Status", "Citizenship/Nationality", "Disability Status", "Date of Interaction", "Is Anonymous"
        ]
        ws.append(headers)

        # Row 2: missing first_name, invalid sex, future DOB
        dob_future = (date.today().replace(year=date.today().year + 1)).strftime('%Y-%m-%d')
        ws.append([
            "123", "", "Doe", "15 - 19", dob_future, "Unknown", "Ward A", "Village B", "Gaborone",
            "", "invalid_email@", "123abc", "", "", date.today().strftime('%Y-%m-%d'), "FALSE"
        ])

        file_obj = BytesIO()
        wb.save(file_obj)
        file_obj.name = 'invalid_template.xlsx'
        file_obj.seek(0)

        response = self.client.post(self.url, {'file': file_obj}, format='multipart')

        # Flatten the error/warning messages
        error_msgs = "\n".join(response.data.get("errors", [])) + "\n" + "\n".join(response.data.get("warnings", []))

        self.assertIn("requires a first name", error_msgs)
        self.assertIn("Sex at column", error_msgs)
        self.assertIn("Date of birth", error_msgs)
    
    def test_kp_dis_fields(self):
        self.client.force_authenticate(user=self.user)

        # Create an Excel file with a single invalid row
        wb = self.create_workbook(self.project.id, self.org.id, include_data=True)  
        ws = wb['Data']

        headers = [
            "ID/Passport Number", "First Name", "Last Name", "Age Range", "Date of Birth", "Sex",
            "Ward", "Village", "District", "Citizenship/Nationality", "Email Address", "Phone Number",
            "Key Population Status", "Disability Status", "Date of Interaction", "Is Anonymous"
        ]
        ws.append(headers)

        # Row 2: missing first_name, invalid sex, future DOB
        dob = (date(2000, 6, 1).strftime('%Y-%m-%d'))
        doi = (date(2025, 6, 1).strftime('%Y-%m-%d'))
        ws.append([
            "123", "Goolius", "Doe", "15 - 19", dob_future, "Unknown", "Ward A", "Village B", "Gaborone",
            "Motswana", "invalid_email@", "75123456", "", "", doi, "FALSE", 
        ])

        file_obj = BytesIO()
        wb.save(file_obj)
        file_obj.name = 'invalid_template.xlsx'
        file_obj.seek(0)

        response = self.client.post(self.url, {'file': file_obj}, format='multipart')

        # Flatten the error/warning messages
        error_msgs = "\n".join(response.data.get("errors", [])) + "\n" + "\n".join(response.data.get("warnings", []))

        self.assertIn("requires a first name", error_msgs)
        self.assertIn("Sex at column", error_msgs)
        self.assertIn("Date of birth", error_msgs)