from django.shortcuts import render, redirect
from django.http import JsonResponse
from django.db.models import Q
from rest_framework import filters
from rest_framework import status
from dateutil.parser import parse as parse_date
from rest_framework.response import Response
from users.restrictviewset import RoleRestrictedViewSet
from rest_framework.permissions import IsAuthenticated
from rest_framework.exceptions import PermissionDenied
from rest_framework.permissions import IsAuthenticated
from dateutil.parser import parse as parse_date
from datetime import date, timedelta
from rest_framework.decorators import action
from rest_framework.exceptions import ValidationError
from rest_framework.filters import OrderingFilter
from rest_framework import serializers
from rest_framework.filters import SearchFilter
from openpyxl.utils.datetime import from_excel
from django.db.models import Q
from django.utils.timezone import now
from openpyxl import Workbook
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.worksheet.protection import SheetProtection
from openpyxl import load_workbook
from io import BytesIO
from django.http import HttpResponse
import os
import re
from django.conf import settings
import string
from itertools import product
from django.db import transaction
import traceback

from datetime import datetime, date
today = date.today().isoformat()
from projects.models import Task
from respondents.models import Respondent, Interaction, Pregnancy, HIVStatus, KeyPopulation, DisabilityType, RespondentAttributeType
from respondents.serializers import RespondentSerializer, RespondentListSerializer, InteractionSerializer, SensitiveInfoSerializer
from indicators.models import IndicatorSubcategory



class RespondentViewSet(RoleRestrictedViewSet):
    permission_classes = [IsAuthenticated]
    filter_backends = [filters.SearchFilter, OrderingFilter]
    ordering_fields = ['last_name', 'first_name', 'village', 'district']
    search_fields = ['first_name', 'last_name', 'uuid', 'comments', 'village'] 
    filterset_fields = ['sex', 'age_range', 'district']
    queryset = Respondent.objects.all()
    serializer_class = RespondentSerializer
    def get_queryset(self):
        queryset = Respondent.objects.all()
        sex = self.request.query_params.get('sex')
        district = self.request.query_params.get('district')
        age_range = self.request.query_params.get('age_range')

        if sex:
            queryset = queryset.filter(sex=sex)
        if district:
            queryset = queryset.filter(district=district)
        if age_range:
            queryset = queryset.filter(age_range=age_range)

        return queryset
    
    def get_serializer_class(self):
        if self.action == 'list':
            return RespondentListSerializer
        else:
            return RespondentSerializer
    
    def perform_create(self, serializer):
        serializer.save(created_by=self.request.user)

    def perform_update(self, serializer):
        serializer.save(updated_by=self.request.user)

    def destroy(self, request, *args, **kwargs):
        user = request.user
        instance = self.get_object()

        # Prevent deletion if respondent has interactions
        if Interaction.objects.filter(respondent_id=instance.id).exists():
            return Response(
                {
                    "detail": (
                        "You cannot delete a respondent that has interactions associated with them. "
                        "If this respondent has requested data removal, consider marking them as anonymous."
                    )
                },
                status=status.HTTP_409_CONFLICT
            )

        # Permission check: only admin can delete
        if user.role != 'admin':
            return Response(
                {"detail": "You do not have permission to delete this respondent."},
                status=status.HTTP_403_FORBIDDEN
            )

        # Perform deletion
        self.perform_destroy(instance)
        return Response(status=status.HTTP_204_NO_CONTENT)


    @action(detail=False, methods=['get'], url_path='meta')
    def filter_options(self, request):
        districts = [district for district, _ in Respondent.District.choices]
        district_labels = [d.label for d in Respondent.District]
        sexs = [sex for sex, _ in Respondent.Sex.choices]
        sex_labels = [choice.label for choice in Respondent.Sex]
        age_ranges = [ar for ar, _ in Respondent.AgeRanges.choices]
        age_range_labels = [choice.label for choice in Respondent.AgeRanges]
        kp_types = [kp for kp, _ in KeyPopulation.KeyPopulations.choices]
        kp_type_labels = [choice.label for choice in KeyPopulation.KeyPopulations]
        dis_types = [dis for dis, _ in DisabilityType.DisabilityTypes.choices]
        dis_labels = [dis.label for dis in DisabilityType.DisabilityTypes]
        auto_attr = [RespondentAttributeType.Attributes.PLWHIV, RespondentAttributeType.Attributes.KP, RespondentAttributeType.Attributes.PWD]
        special_attributes = [attr for attr, _ in RespondentAttributeType.Attributes.choices if attr not in auto_attr]
        special_attribute_labels = [attr.label for attr in RespondentAttributeType.Attributes if attr not in auto_attr]
        return Response({
            'districts': districts,
            'district_labels': district_labels,
            'sexs': sexs,
            'sex_labels': sex_labels,
            'age_ranges': age_ranges,
            'age_range_labels': age_range_labels,
            'kp_types': kp_types,
            'kp_type_labels': kp_type_labels,
            'disability_types': dis_types,
            'disability_type_labels': dis_labels,
            'special_attributes': special_attributes,
            'special_attribute_labels': special_attribute_labels
        })
    
    #we should write some tests for this at some point
    @action(detail=True, methods=['get', 'post', 'patch'], url_path='sensitive-info')
    def sensitive_info(self, request, pk=None):
        respondent = self.get_object()
        if request.method == 'GET':
            serializer = SensitiveInfoSerializer(respondent)
            return Response(serializer.data)

        elif request.method == 'POST' or request.method == 'PATCH':
            if request.user.role == 'client':
                raise PermissionDenied('You do not have permission to perform this action.')
            serializer = SensitiveInfoSerializer(respondent, data=request.data, partial=True)
            serializer.is_valid(raise_exception=True)
            serializer.save(updated_by=request.user)
            return Response(serializer.data, status=status.HTTP_200_OK)
        
    @action(detail=False, methods=['post'], url_path='bulk')
    def bulk_upload(self, request):
        if request.user.role == 'client':
                raise PermissionDenied('You do not have permission to perform this action.')
        data = request.data
        if not isinstance(data, list):
            return Response({"detail": "Expected a list of respondents."}, status=400)

        created_ids = []
        local_ids = []
        errors = []
        for i, item in enumerate(data):
            try:
                sensitive_data = item.pop("sensitive_info", {})
                interactions = item.pop("interactions", [])

                id_no = item.get('id_no')
                existing = None
                if id_no is not None:
                    existing = Respondent.objects.filter(id_no=id_no).first()
                if existing:
                    respondent_serializer = RespondentSerializer(
                        existing, data=item, context={'request': request}, partial=True
                    )
                else:
                    respondent_serializer = RespondentSerializer(
                        data=item, context={'request': request}
                    )
                respondent_serializer.is_valid(raise_exception=True)
                respondent = respondent_serializer.save(created_by=request.user)

                # Save sensitive info if provided
                if sensitive_data:
                    sensitive_serializer = SensitiveInfoSerializer(
                        instance=respondent,
                        data=sensitive_data,
                        partial=True,
                        context={'request': request}
                    )
                    sensitive_serializer.is_valid(raise_exception=True)
                    sensitive_serializer.save(updated_by=request.user)

                # Save interactions
                with transaction.atomic():
                    for interaction in interactions:
                        try:
                            interaction_date = interaction.get('interaction_date')
                            if not interaction_date:
                                raise ValidationError({'interaction_date': 'Interaction date is required'})
                            interaction_location = interaction.get('interaction_location') or None
                            subcats = interaction.get("subcategories_data", [])
                            task_id = interaction.get("task")
                            numeric_component = interaction.get('numeric_component')
                            try:
                                task_instance = Task.objects.get(id=task_id)
                            except Task.DoesNotExist:
                                raise ValidationError({"task": f"Task with ID {task_id} not found"})
                            lookup_fields = {
                                'respondent': respondent,
                                'interaction_date': interaction_date,
                                'task': task_id,
                            }
                            # Try to fetch the existing interaction
                            instance = Interaction.objects.filter(**lookup_fields).first()

                            # Pass instance to serializer if it exists (update), otherwise it will create
                            serializer = InteractionSerializer(
                                instance=instance,
                                data={
                                    'respondent': respondent.id,
                                    'interaction_date': interaction_date,
                                    'interaction_location': interaction_location,
                                    'task': task_id,
                                    'numeric_component': numeric_component,
                                    'subcategories_data': subcats,
                                    'comments': '',
                                },
                                context={'request': request, 'respondent': respondent}
                            )
                            serializer.is_valid(raise_exception=True)
                            serializer.save()
                        except Exception as inter_err:
                            errors.append({
                                'respondent': respondent.id,
                                'interaction_error': str(inter_err),
                                'interaction_traceback': traceback.format_exc(),
                                'interaction_data': interaction
                            })
            
                local_ids.append(item.get('local_id'))
                created_ids.append(respondent.id)

            except ValidationError as ve:
                errors.append({
                    'index': i,
                    'error': 'Validation error',
                    'details': ve.detail,
                    'data': item
                })
            except Exception as e:
                errors.append({
                    'index': i,
                    'error': str(e),
                    'data': item
                })
        print(errors)
        return Response({
            "created_ids": created_ids,
            "local_ids": local_ids,
            "errors": errors
        }, status=status.HTTP_207_MULTI_STATUS if errors else status.HTTP_201_CREATED)
    
class InteractionViewSet(RoleRestrictedViewSet):
    permission_classes = [IsAuthenticated]
    queryset = Interaction.objects.all()
    serializer_class = InteractionSerializer
    ordering_fields = ['-interaction_date']
    filter_backends = [SearchFilter]
    filterset_fields = ['task', 'respondent', 'interaction_date']
    search_fields = ['respondent__uuid', 'respondent__first_name', 'respondent__last_name', 
                     'task__indicator__code', 'task__indicator__name', 'task__organization__name'] 
    def get_queryset(self):
        queryset = super().get_queryset()
        respondent = self.request.query_params.get('respondent')
        user = self.request.user
        role = getattr(user, 'role', None)
        org = getattr(user, 'organization', None)

        if respondent:
            queryset = queryset.filter(respondent__id=respondent)

        start = self.request.query_params.get('start')
        end = self.request.query_params.get('end')
        if start:
            queryset = queryset.filter(interaction_date__gte=start)
        if end:
            queryset = queryset.filter(interaction_date__lte=end)
        return queryset

    
    def destroy(self, request, *args, **kwargs):
        user = request.user  # consistent access
        instance = self.get_object()

        if user.role != 'admin':
            return Response(
                {
                    "detail": "You do not have permission to delete this interaction."
                },
                status=status.HTTP_403_FORBIDDEN 
            )
        if Interaction.objects.filter(respondent=instance.respondent, task__indicator__prerequisite=instance.task.indicator).exists():
            return Response(
                {
                    "detail": "Another interaction is relying on this as a prerequisite interaction. Please delete that interaction first."
                },
                status=status.HTTP_409_CONFLICT
            )

        self.perform_destroy(instance)
        return Response(status=status.HTTP_204_NO_CONTENT)
    
    @action(detail=False, methods=['get'], url_path='flagged')
    def get_flagged(self, request):
        user = request.user
        role = user.role
        org = user.organization

        # Start with flagged interactions
        queryset = Interaction.objects.filter(flagged=True)

        # Role-based filtering
        if role == 'client':
            raise PermissionDenied('You do not have permission to view this page.')
        elif role in ['meofficer', 'manager']:
            queryset = queryset.filter(
                Q(task__organization=org) |
                Q(task__organization__parent_organization=org)
            )
        elif role == 'data_collector':
            queryset = queryset.filter(created_by=user)

        project_id = request.query_params.get('project')
        organization_id = request.query_params.get('organization')
        indicator_id = request.query_params.get('indicator')
        
        if project_id:
            queryset = queryset.filter(task__project__id = project_id)
        if organization_id:
            queryset = queryset.filter(task__organization__id = organization_id)
        if indicator_id:
            queryset = queryset.filter(task__indicator__id = indicator_id)

        # Search filter (e.g., by respondent name or ID or comment)
        search_term = request.query_params.get('search')
        if search_term:
            queryset = queryset.filter(
                Q(respondent__first_name__icontains=search_term) |
                Q(respondent__last_name__icontains=search_term) |
                Q(respondent__village__icontains=search_term) |
                Q(respondent__uuid__icontains=search_term) |
                Q(comments__icontains=search_term) |
                Q(task__indicator__name__icontains=search_term) |
                Q(task__indicator__code__icontains=search_term) |
                Q(task__organization__name__icontains=search_term) |
                Q(task__project__name__icontains=search_term)
            )

        # Pagination
        page = self.paginate_queryset(queryset)
        if page is not None:
            serializer = InteractionSerializer(page, many=True, context={'request': request})
            return self.get_paginated_response(serializer.data)

        # No pagination fallback
        serializer = InteractionSerializer(queryset, many=True, context={'request': request})
        return Response(serializer.data)
    
    @action(detail=False, methods=['post', 'patch'], url_path='batch')
    def batch_create(self, request):
        if request.user.role == 'client':
                raise PermissionDenied('You do not have permission to perform this action.')
        respondent_id = request.data.get('respondent')
        tasks = request.data.get('tasks', [])
        top_level_date = request.data.get('interaction_date')
        top_level_location = request.data.get('interaction_location')
        if not respondent_id or not tasks:
            return Response({'error': 'Missing respondent or tasks'}, status=status.HTTP_400_BAD_REQUEST)

        try:
            respondent = Respondent.objects.get(id=respondent_id)

        except Respondent.DoesNotExist:
            return Response({'error': 'Respondent not found'}, status=status.HTTP_404_NOT_FOUND)

        created = []
        for i, task in enumerate(tasks):
            print(task)
            task_date = task.get('interaction_date') or top_level_date
            task_location = task.get('interaction_location') or top_level_location
            if not task_date:
                return Response({'error': f'Missing interaction_date for task index {i}'}, status=status.HTTP_400_BAD_REQUEST)

            if not all(k in task for k in ['task']):
                return Response({'error': f'Missing required task fields at index {i}'}, status=status.HTTP_400_BAD_REQUEST)
        
            serializer = self.get_serializer(data={
                'respondent': respondent_id,
                'interaction_date': task_date,
                'interaction_location': task_location,
                'task': task['task'],
                'numeric_component': task.get('numeric_component'),
                'subcategories_data': task.get('subcategories_data', []),
                'comments': task.get('comments', ''),
            }, context={'request': request, 'respondent': respondent})

            serializer.is_valid(raise_exception=True)
            serializer.save()
            created.append(serializer.data)

        return Response(created, status=status.HTTP_201_CREATED)






    @staticmethod
    def excel_columns():
        for size in range(1, 3):  # A to ZZ
            for letters in product(string.ascii_uppercase, repeat=size):
                yield ''.join(letters)

    @action(detail=False, methods=['get'], url_path='template')
    def get_template(self, request):
        from projects.models import Task
        from organizations.models import Organization
        user=request.user
        if not user.role in ['meofficer', 'manager', 'admin']:
            raise PermissionDenied('You do not have permission to access templates.')
        project_id = request.GET.get('project')
        org_id = request.GET.get('organization')
        if not project_id or not org_id:
            raise serializers.ValidationError('Template requires a project and organization.')
        valid_orgs = Organization.objects.filter(Q(parent_organization=user.organization) | Q(id=user.organization.id))
        if user.role != 'admin' and not valid_orgs.filter(id=org_id).exists():
            raise PermissionDenied('You do not have permission to access this template.')
        
        district_labels = [choice.label for choice in Respondent.District]
        sex_labels = [choice.label for choice in Respondent.Sex]
        age_range_labels = [choice.label for choice in Respondent.AgeRanges]
        sex_labels = [choice.label for choice in Respondent.Sex]
        kp_type_labels = [choice.label for choice in KeyPopulation.KeyPopulations]
        dis_labels = [dis.label for dis in DisabilityType.DisabilityTypes]

        headers = []
        for field in Respondent._meta.get_fields():
            if field.auto_created:
                continue
            if field.name in ['uuid', 'created_by', 'created_at', 'updated_at', 'updated_by']:
                continue
            if hasattr(field, 'verbose_name'):
                verbose = field.verbose_name
            else:
                verbose = field.name

            field_info = {'header': verbose or field.name}
            if field.many_to_many:
                field_info['multiple'] = True
            else:
                field_info['multiple'] = False

            if field.name == 'is_anonymous':
                field_info['options'] = ['TRUE', 'FALSE']
            elif field.name == 'district':
                field_info['options'] = district_labels
            elif field.name == 'sex':
                field_info['options'] = sex_labels
            elif field.name == 'age_range':
                field_info['options'] = age_range_labels
            elif field.name == 'kp_status':
                field_info['options'] = kp_type_labels
            elif field.name == 'disability_status':
                field_info['options'] = dis_labels
            else:
                field_info['options'] = []
            headers.append(field_info)
        
        headers.append({'header': 'HIV Status', 'options': ['HIV Positive', 'HIV Negative'], 'multiple': False})
        headers.append({'header': 'Date Positive', 'options': [], 'multiple': False})
        headers.append({'header': 'Pregnant', 'options': ['Yes', 'No'], 'multiple': False})
        headers.append({'header': 'Date of Interaction', 'options': [], 'multiple': False})
        headers.append({'header': 'Interaction Location', 'options': [], 'multiple': False})
        tasks = Task.objects.filter(organization__id=org_id, project__id=project_id).order_by('indicator__code')
        if not tasks:
            raise serializers.ValidationError('There are no tasks associated with this project for your organization.')
        
        project_name = tasks[0].project.name.replace(' ', '').replace('/', '-')[:25]
        for task in tasks:
            header = task.indicator.code + ': ' + task.indicator.name
            if task.indicator.require_numeric:
                header = header + ' (Requires a Number)'
            categories = []
            subcats = task.indicator.subcategories.all()
            if subcats.exists():
                for cat in subcats:
                    categories.append(cat.name)
            elif not task.indicator.require_numeric:
                categories = ['Yes', 'No']
            headers.append({'header': header, 'options': categories, 'multiple': True})
        
        template_path = os.path.join(settings.BASE_DIR, 'respondents', 'static', 'respondents', 'upload-template.xlsx')
        wb = load_workbook(template_path)
        ws = wb.create_sheet('Data') #add a protection to this name or something
        options_sheet = wb.create_sheet('DropdownOptions')
        cols = list(self.excel_columns())
        for c, header in enumerate(headers, 1):
            header_text = header['header']
            ws.cell(row=1, column=c, value=header_text)
            if len(header['options']) > 0:
                col = cols[c-1]
                for r, option in enumerate(header['options'], 1):
                    options_sheet[f'{col}{r}'] = str(option)
                dv = DataValidation(type="list", formula1=f"=DropdownOptions!${col}$1:${col}${len(header['options'])}", allow_blank=True)
                dv.add(f'{cols[c-1]}2:{cols[c-1]}1000')
                ws.add_data_validation(dv)
            else:
                continue
        options_sheet.sheet_state = 'hidden'

        metadata_sheet = wb.create_sheet("Metadata")
        metadata_sheet["A1"] = "project_id"
        metadata_sheet["B1"] = project_id
        metadata_sheet["A2"] = "organization_id"
        metadata_sheet["B2"] = org_id

        metadata_sheet.sheet_state = 'hidden'
        metadata_sheet.protection.sheet = True
        metadata_sheet.protection.password = 'xQvzLit1@SS69'

        output = BytesIO()
        wb.save(output)
        output.seek(0)
        response = HttpResponse(output.read(), content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        filename = f'{project_name}_{date.today().strftime("%Y-%m-%d")}.xlsx'
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        return response

    @action(detail=False, methods=['POST'], url_path='upload')
    def post_template(self, request):
        from projects.models import Task, Project
        from organizations.models import Organization
        errors = []
        warnings = []
        user = request.user
        
        if user.role not in ['admin', 'meofficer', 'manager']:
            raise PermissionDenied('You do not have permission to access templates.')

        uploaded_file = request.FILES.get('file')
        if not uploaded_file:
            return Response({"detail": "No file was uploaded."}, status=status.HTTP_400_BAD_REQUEST)
        if not uploaded_file.name.endswith('.xlsx'):
            return Response({"detail": "Uploaded file must be an .xlsx Excel file."}, status=status.HTTP_400_BAD_REQUEST)
        
        try:
            wb = load_workbook(filename=uploaded_file)
            ws = wb['Metadata']
        except Exception:
            raise serializers.ValidationError("Unable to read 'Metadata' sheet. Please check the template.")

        try:
            project_id = int(ws['B1'].value)
            org_id = int(ws['B2'].value)
        except (TypeError, ValueError):
            raise serializers.ValidationError("Project ID and Organization ID must be numeric.")
        if not project_id or not org_id:
            raise serializers.ValidationError("Template requires both a valid project and organization ID.")

        project = Project.objects.filter(id=project_id).first()
        # 4. Organization permission check
    
        if user.role != 'admin':
            valid_orgs = Organization.objects.filter(Q(parent_organization=user.organization) | Q(id=user.organization_id))
            if not valid_orgs.filter(id=org_id).exists():
                raise PermissionDenied('You do not have permission to access this template.')

        ws = wb['Data'] 
        headers = {}
        for row in ws.iter_rows(min_row=1, max_row=1):
            for cell in row:
                if cell.value:
                    header_name = str(cell.value).strip()
                    headers[header_name] = {
                        'column': cell.column,
                        'options': [],
                        'multiple': False
                    }

        #for simplicity/verification, we're treating everything as lowercase no spaces
        district_labels = [choice.label.lower().replace(' ', '') for choice in Respondent.District]
        sex_labels = [choice.label.lower().replace(' ', '') for choice in Respondent.Sex]
        age_range_labels = [choice.label.lower().replace(' ', '')  for choice in Respondent.AgeRanges]
        kp_type_labels = [choice.label.lower().replace(' ', '')  for choice in KeyPopulation.KeyPopulations]
        dis_labels = [dis.label.lower().replace(' ', '')  for dis in DisabilityType.DisabilityTypes]
        
        def get_verbose(field_name):
            return Respondent._meta.get_field(field_name).verbose_name
        
        def expect_column(field_name, options=None, multiple=False):
            verbose = get_verbose(field_name)
            if verbose in headers:
                if options:
                    headers[verbose]['options'] = options
                headers[verbose]['multiple'] = multiple
            else:
                errors.append(f"Template is missing {verbose} column.")
        
        expect_column('id_no')
        expect_column('first_name')
        expect_column('last_name')
        expect_column('age_range', options=age_range_labels)
        expect_column('dob')
        expect_column('sex', options = sex_labels)
        expect_column('ward')
        expect_column('village')
        expect_column('district', options=district_labels)
        expect_column('citizenship')
        expect_column('email')
        expect_column('phone_number')
        expect_column('kp_status', options=kp_type_labels, multiple=True)
        expect_column('disability_status', options=dis_labels, multiple=True) 

        if not 'Date of Interaction' in headers:
            errors.append('Template is missing Date of Interaction column.')
        if not 'Interaction Location' in headers:
            errors.append('Template is missing Interaction Location column.')
        if not 'HIV Status' in headers:
            errors.append('Template is missing HIV Status column.')
        if not 'Date Positive' in headers:
            errors.append('Template is missing Date Positive column.')
        if not 'Pregnant' in headers:
            errors.append('Template is missing Pregnant column.')
        
        tasks = Task.objects.filter(organization__id=org_id, project__id=project_id).order_by('indicator__code')
        if not tasks:
           errors.append('No tasks associted with project.')
           return Response({'errors': errors, 'warnings': warnings,  }, status=status.HTTP_400_BAD_REQUEST)
        
        for task in tasks:
            header = task.indicator.code + ': ' + task.indicator.name
            if task.indicator.require_numeric:
                header = header + ' (Requires a Number)'
            categories = []
            subcats = task.indicator.subcategories.all()
            if subcats.exists():
                categories = [cat.name.lower().replace(' ', '') for cat in subcats]
            if header in headers:
                headers[header]['options'] = categories
                headers[header]['multiple'] = True
            else:
                warnings.append(f'Task {header} is missing from this template. It may be invalid or out of date.')

        def get_cell_value(row, field_name):
            header = headers.get(get_verbose(field_name))
            if header:
                return row[header['column'] - 1]
            return None

        def get_column(field_name):
            header = headers.get(get_verbose(field_name))
            if header:
                return header['column']
            return None

        def get_task_value(row, task):
            header_name = task.indicator.code + ': ' + task.indicator.name
            if task.indicator.require_numeric:
                header_name += ' (Requires a Number)'
            header = headers.get(header_name)
            if header:
                return row[header['column'] - 1]
            return None

        def get_task_column(task):
            header_name = task.indicator.code + ': ' + task.indicator.name
            if task.indicator.require_numeric:
                header_name += ' (Requires a Number)'
            header = headers.get(header_name)
            if header:
                return header['column']
            return None

        def get_task_options(task):
            header_name = task.indicator.code + ': ' + task.indicator.name
            if task.indicator.require_numeric:
                header_name += ' (Requires a Number)'
            header = headers.get(header_name)
            if header:
                return header['options']
            return []

        def get_options(field_name):
            header = headers.get(get_verbose(field_name))
            if header:
                return header['options']
            return []

        
        def valid_excel_date(value):
            if value is None:
                return None
            # Already a Python date or datetime
            if isinstance(value, datetime):
                value = value.date()
            if isinstance(value, date):
                if value > date.today():
                    return None
                return value
            # Try ISO string
            try:
                parsed = date.fromisoformat(value)
                if parsed > date.today():
                    return None
                return parsed
            except (ValueError, TypeError):
                pass
            # Try Excel serial number (e.g., 45000 or '45000')
            try:
                numeric_value = float(value)
                converted = from_excel(numeric_value)
                if isinstance(converted, datetime):
                    converted = converted.date()
                if converted > date.today():
                    return None
                return converted
            except (ValueError, TypeError):
                pass
            try:
                parsed = parse_date(value, dayfirst=True).date()
                if parsed > date.today():
                    return None
                return parsed
            except (ValueError, TypeError):
                pass
            try:
                parsed = parse_date(value, dayfirst=False).date()
                if parsed > date.today():
                    return None
                return parsed
            except (ValueError, TypeError):
                pass
            return None

        def is_email(value):
            pattern = r'^[\w\.-]+@[\w\.-]+\.\w+$'
            return bool(re.match(pattern, value))

        def is_phone_number(value):
            pattern = r'^\+?[\d\s\-\(\)]{7,20}$'
            return bool(re.fullmatch(pattern, value))
        
        if len(errors) > 0:
            return Response({'errors': errors, 'warnings': warnings,  }, status=status.HTTP_400_BAD_REQUEST)
        
        def is_truthy(value):
            if isinstance(value, bool):
                return value
            if isinstance(value, (int, float)):
                return value == 1
            if isinstance(value, str):
                return value.strip().lower() in ['true', 'yes', '1']
            return False

        for i, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
            respondent = None
            row_errors = []
            row_warnings = []
            anon = is_truthy(get_cell_value(row, 'is_anonymous'))
            id_no = get_cell_value(row, 'id_no') or None
            if id_no and not anon:
                respondent = Respondent.objects.filter(id_no=id_no).first()
                if respondent:
                    row_warnings.append(f"Respondent {respondent} at column: {get_column('id_no')}, row: {i} already exists.")
            
            first_name = get_cell_value(row, 'first_name') or None
            last_name = get_cell_value(row, 'last_name') or None

            if not anon and not first_name:
                row_errors.append(f"Respondent at column: {get_column('first_name')}, row: {i} requires a first name.")
            if not anon and not last_name:
                row_errors.append(f"Respondent at column: {get_column('last_name')}, row: {i} requires a last name.")

            dob = get_cell_value(row, 'dob') or None
            if anon:
                dob = None
            else:
                if dob:
                    parsed = valid_excel_date(dob)
                    if not parsed:
                        row_errors.append(f"Date of birth {dob} at column: {get_column('dob')}, row: {i} is invalid. Double check the format and make sure that it is not in the future")
                    dob = parsed
                else:
                    row_errors.append(f"Date of birth at column: {get_column('dob')}, row: {i} is required for non-anonymous respondents.")


            
            sex = get_cell_value(row, 'sex')
            if sex:
                sex=sex.lower().replace(' ', '')
                if not sex in get_options('sex'):
                    row_errors.append(f"Sex at column: {get_column('sex')}, row: {i} is not a valid choice.")
            else:
                row_errors.append(f"Sex at column: {get_column('sex')}, row: {i} is required.")

            age_range = get_cell_value(row, 'age_range')
            if anon:
                if age_range:
                    age_range = age_range.lower().replace(' ', '')
                    if not age_range in get_options('age_range'):
                        row_errors.append(f"Age Range value at column: {get_column('age_range')}, row: {i} is not a valid choice.")
                else:
                    row_errors.append(f"Age range at column: {get_column('sex')}, row: {i} is required for anonymous respondents.")
            ward = get_cell_value(row, 'ward') or None
            
            village = get_cell_value(row, 'village') or None
            if not village:
                row_errors.append(f"Village at column: {get_column('village')}, row: {i} is required for all respondents.")

            district = get_cell_value(row, 'district') or None
            if district:
                district = district.lower().replace(' ', '')
                if not district in get_options('district'):
                    row_errors.append(f"District at column: {get_column('district')}, row: {i} is not a valid choice.")
            else:
                row_errors.append(f"District at column: {get_column('district')}, row: {i} is required.")
            
            citizenship = get_cell_value(row, 'citizenship') or None
            if not citizenship:
                citizenship = 'Motswana'
                row_errors.append(f"Citizenship at column: {get_column('citizenship')}, row: {i} is required for all respondents. This value will default to Motswana. If this is incorrect, please check this field again.")
            
            email = get_cell_value(row, 'email') or None
            if email and not is_email(email):
                row_warnings.append(f"Email at column: {get_column('email')}, row: {i} is not a valid choice.")
                email = None
            phone_number = get_cell_value(row, 'phone_number') or None
            if phone_number and not is_phone_number(phone_number):
                row_warnings.append(f"Phone Number at column: {get_column('phone_number')}, row: {i} is not a valid choice.")
                phone_number = None
            comments = get_cell_value(row, 'comments') or None
            
            if len(row_errors) > 0:
                row_errors.append("This respondent and their interactions will not be saved until these errors are fixed")
                errors.extend(row_errors)
                warnings.extend(row_warnings)
                continue

            def get_choice_key_from_label(choices, label):
                for key, value in choices:
                    if value.lower().replace(' ', '') == label.lower().replace(' ', ''):
                        return key
                return None
            
            sex = get_choice_key_from_label(Respondent.Sex.choices, sex)
            district = get_choice_key_from_label(Respondent.District.choices, district)
            if age_range and not dob:
                age_range = get_choice_key_from_label(Respondent.AgeRanges.choices, age_range)
            

            if not respondent:
                respondent = Respondent.objects.create(
                    is_anonymous = anon,
                    id_no = id_no,
                    first_name = first_name,
                    last_name = last_name,
                    ward = ward,
                    village = village,
                    district = district,
                    sex = sex,
                    dob = dob,
                    age_range = age_range,
                    citizenship = citizenship,
                    email = email,
                    phone_number = phone_number,
                    comments = comments,
                    created_by=user
                )

            print(respondent)
            
                
            kp_status_names_raw = get_cell_value(row, 'kp_status') or ''
            if kp_status_names_raw:
                # Clean and split
                cleaned = kp_status_names_raw.replace(' ', '').lower()
                input_kp_names = set(re.split(r'[,:;]', cleaned))

                valid_labels = get_options('kp_status')
                valid_lookup = {label.replace(' ', '').lower(): label for label in valid_labels}

                matched = [name for name in input_kp_names if name in valid_lookup]
                invalid_kp = input_kp_names - set(matched)

                if invalid_kp:
                    row_warnings.append(
                        f"Invalid key population statuses at row {i}: {', '.join(invalid_kp)}"
                    )

                kp_types = []
                for cleaned_name in matched:
                    key = get_choice_key_from_label(KeyPopulation.KeyPopulations.choices, cleaned_name)
                    if not key:
                        continue
                    kp, _ = KeyPopulation.objects.get_or_create(name=key)
                    kp_types.append(kp)

                if kp_types:
                    respondent.kp_status.set(kp_types)
            

            disability_status_names_raw = get_cell_value(row, 'disability_status') or ''
            if disability_status_names_raw:
                # Clean and split
                cleaned = disability_status_names_raw.replace(' ', '').lower()
                input_disability_names = set(re.split(r'[,:;]', cleaned))

                valid_labels = get_options('disability_status')
                valid_lookup = {label.replace(' ', '').lower(): label for label in valid_labels}

                matched = [name for name in input_disability_names if name in valid_lookup]
                invalid_disability = input_disability_names - set(matched)

                if invalid_disability:
                    row_warnings.append(
                        f"Invalid disability statuses at row {i}: {', '.join(invalid_disability)}"
                    )
                disability_types = []
                for cleaned_name in matched:
                    key = get_choice_key_from_label(DisabilityType.DisabilityTypes.choices, cleaned_name)
                    if not key:
                        continue
                    dis, _ = DisabilityType.objects.get_or_create(name=key)
                    disability_types.append(dis)

                if disability_types:
                    respondent.disability_status.set(disability_types)
            
            hs_col = headers['HIV Status']['column']-1 
            hiv_status = row[hs_col] if len(row) > hs_col else None
            if hiv_status:
                if hiv_status.lower().replace(' ', '') == 'yes':
                    hiv_status = True
                    dp_col = headers['Date Positive']['column']-1 
                    date_positive = row[dp_col] if len(row) > dp_col else None
                    parsed = valid_excel_date(date_positive)
                    if not parsed:
                        row_errors.append(
                            f"Date positive '{date_positive}' at column: {dp_col}, row: {i} is invalid."
                            "Double check the format and make sure that it is not in the future."
                        )
                    else:
                        date_positive = parsed
                    HIVStatus.objects.create(respondent=respondent, hiv_positive=hiv_status, date_positive=date_positive)
            preg_col = headers['Pregnant']['column']-1 
            pregnancy = row[preg_col] if len(row) > preg_col else None
            if pregnancy:
                if pregnancy.lower().replace(' ', '') == 'yes':
                    Pregnancy.objects.create(respondent=respondent, is_pregnant=True, term_began=date.today())
                elif pregnancy.lower().replace(' ', '') == 'no' and Pregnancy.objects.filter(respondent=respondent, is_pregnant=True).exists():
                    pregnancy = Pregnancy.objects.filter(respondent=respondent, is_pregnant=True).first()
                    pregnancy.is_pregnant = False
                    pregnancy.term_ended = date.today()
                    pregnancy.save()


            doi_col = headers['Date of Interaction']['column']-1 
            interaction_date = row[doi_col] if len(row) > doi_col else None

            if interaction_date:
                parsed = valid_excel_date(interaction_date)
                if not parsed:
                    row_errors.append(
                        f"Date of interaction '{interaction_date}' at column: {doi_col}, row: {i} is invalid. "
                        "Double check the format and make sure that it is not in the future."
                    )
                else:
                    interaction_date = parsed
                    if not (project.start <= interaction_date <= project.end):
                        row_errors.append(
                            f"Date of interaction '{interaction_date}' at column: {doi_col}, row: {i} is outside "
                            "of the range of this project."
                        )
            else:
                row_errors.append(
                    f"Date of interaction  at column: {doi_col}, row: {i} is required. "
                )
            loc_col = headers['Interaction Location']['column']-1 
            interaction_location = row[loc_col] if len(row) > loc_col else None

            if len(row_errors) > 0:
                row_errors.append(f"This respondent has been saved, but none of their interactions will until this date is fixed.")
                errors.extend(row_errors)
                warnings.extend(row_warnings)
                continue

            def topological_sort(tasks):
                from collections import defaultdict, deque

                graph = defaultdict(list)
                in_degree = defaultdict(int)

                for task in tasks:
                    if task.indicator.prerequisite:
                        graph[task.indicator.prerequisite.id].append(task.indicator.id)
                        in_degree[task.indicator.id] += 1
                    else:
                        in_degree[task.indicator.id] += 0

                id_map = {task.indicator.id: task for task in tasks}

                queue = deque([id for id in in_degree if in_degree[id] == 0])
                sorted_ids = []

                while queue:
                    current = queue.popleft()
                    sorted_ids.append(current)
                    for dependent in graph[current]:
                        in_degree[dependent] -= 1
                        if in_degree[dependent] == 0:
                            queue.append(dependent)

                if len(sorted_ids) != len(tasks):
                    raise Exception("Cycle detected in prerequisites")
                
                return [id_map[i] for i in sorted_ids]

            for task in topological_sort(tasks):
                col = get_task_column(task)
                val = str(get_task_value(row, task))
                val = val.lower().replace(' ', '')
                if val in ['', 'no', 'none', 'na', 'n/a', 'false', 'unsure', 'maybe']:
                    continue

                if isinstance(val, str):
                    val = val.lower().replace(' ', '')

                valid_subcats = []
                numeric_component = None
                if task.indicator.subcategories.exists():
                    val = val.split(',') if val else []
                    subcats = []
                    for v in val:
                        v = v.strip().lower()
                        if task.indicator.require_numeric:
                            if not ':' in v:
                                row_errors.append(f'Task {task.indicator.name} at column: {col}, row {i} requires a number (make sure that you have the category name and number seperated by a colon, for example "Category: 5")')
                                continue
                            v = v.split(':')
                            subcats.append({'slug': v[0], 'numeric_component': v[1]})
                        else:
                            subcats.append({'slug': v})
                    valid_subcats = []
                    valid_slugs = list(task.indicator.subcategories.values_list('slug', flat=True))
                    for cat in subcats:
                        if cat['slug'] in valid_slugs:
                            isc = IndicatorSubcategory.objects.filter(slug=cat['slug']).first()
                            sc_data = {'id': isc.id, 'name': isc.name}
                            if task.indicator.require_numeric:
                                sc_data['numeric_component'] = cat['numeric_component']
                            valid_subcats.append(sc_data)
                    if len(valid_subcats) == 0:
                        row_errors.append(f'Task {task.indicator.name} at column: {col}, row: {i} requires valid subcategories.')
                        continue

                elif task.indicator.require_numeric and isinstance(val, str):
                    try:
                        numeric_component = float(val)
                        if numeric_component < 0:
                            row_warnings.append(f'Number at column: {col}, row: {i} must be greater than 0.')
                            continue
                    except (ValueError, TypeError):
                        row_warnings.append(f'Number at column: {col}, row: {i} is not a valid number.')
                        continue
                    
                if val:
                    lookup_fields = {
                        'respondent': respondent,
                        'interaction_date': interaction_date,
                        'task': task,
                    }

                    # Try to fetch the existing interaction
                    instance = Interaction.objects.filter(**lookup_fields).first()

                    # Pass instance to serializer if it exists (update), otherwise it will create
                    serializer = self.get_serializer(
                        instance=instance,
                        data={
                            'respondent': respondent.id,
                            'interaction_date': interaction_date,
                            'interaction_location': interaction_location,
                            'task': task.id,
                            'numeric_component': numeric_component,
                            'subcategories_data': valid_subcats,
                            'comments': '',
                        },
                        context={'request': request, 'respondent': respondent}
                    )

                    try:
                        serializer.is_valid(raise_exception=True)
                        serializer.save()
                    except ValidationError as e:
                        # Flatten error details for easier reading
                        error_details = serializer.errors
                        for field, msgs in error_details.items():
                            if isinstance(msgs, list):
                                for msg in msgs:
                                    row_errors.append(f"Row {i}, Column {col}, Field '{field}': {msg}")
                            else:
                                errors.append(f"Row {i}, Column {col}, Field '{field}': {msgs}")
            errors.extend(row_errors)
            warnings.extend(row_warnings)
        print('warnings', warnings)
        print('errors', errors)
        return Response({'errors': errors, 'warnings': warnings }, status=status.HTTP_200_OK)