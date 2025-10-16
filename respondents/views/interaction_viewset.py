from django.shortcuts import render, redirect
from django.db import transaction
from django.db.models import Q
from django.shortcuts import get_object_or_404
from django.conf import settings
from django.http import HttpResponse

from rest_framework import filters, status
from rest_framework.response import Response
from rest_framework.permissions import IsAuthenticated
from rest_framework.exceptions import PermissionDenied, ValidationError

from rest_framework.filters import OrderingFilter
from rest_framework import serializers
from rest_framework.filters import SearchFilter
from rest_framework.decorators import action

from dateutil.parser import parse as parse_date
from datetime import date, timedelta, datetime
from django.utils.timezone import now

from openpyxl import Workbook
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.worksheet.protection import SheetProtection
from openpyxl import load_workbook
from io import BytesIO
import os
import re
import pycountry

from users.restrictviewset import RoleRestrictedViewSet

from projects.models import Task, Project, ProjectOrganization
from respondents.models import Respondent, Interaction, HIVStatus, KeyPopulation, DisabilityType, RespondentAttributeType
from respondents.serializers import RespondentSerializer, InteractionSerializer
from respondents.utils import check_event_perm
from respondents.utils_file_upload import excel_columns, valid_excel_date, is_email, is_phone_number, is_truthy
from indicators.models import  Indicator, Option
from events.models import Event, EventOrganization

class InteractionViewSet(RoleRestrictedViewSet):
    '''
    Viewset that manages everything related to viewing/creating interactions (including file uploads)
    '''
    permission_classes = [IsAuthenticated]
    serializer_class = InteractionSerializer
    ordering_fields = ['-interaction_date']
    filter_backends = [SearchFilter]
    filterset_fields = ['task', 'respondent', 'interaction_date']
    search_fields = ['respondent__uuid', 'respondent__first_name', 'respondent__last_name', 
                     'task__indicator__code', 'task__indicator__name', 'task__organization__name'] 
    
    def get_queryset(self):
        '''
        In context, interactions are almost always viewed in the context of a respondent, Interactions are
        visible to all site users. 
        '''
        queryset = Interaction.objects.all()
        #URL params. 
        respondent = self.request.query_params.get('respondent')
        user = self.request.user

        if respondent:
            queryset = queryset.filter(respondent__id=respondent)

        start = self.request.query_params.get('start')
        end = self.request.query_params.get('end')
        if start:
            queryset = queryset.filter(interaction_date__gte=start)
        if end:
            queryset = queryset.filter(interaction_date__lte=end)

        indicator_param = self.request.query_params.get('indicator')
        if indicator_param:
            queryset = queryset.filter(task__indicator_id=indicator_param)
        return queryset
    
    def destroy(self, request, *args, **kwargs):
        '''
        Only admins can delete
        '''
        user = request.user  # consistent access
        instance = self.get_object()

        #only admins can delete
        if user.role != 'admin':
            return Response(
                {
                    "detail": "You do not have permission to delete this interaction."
                },
                status=status.HTTP_403_FORBIDDEN 
            )

        self.perform_destroy(instance)
        return Response(status=status.HTTP_204_NO_CONTENT)


    @action(detail=False, methods=['post'], url_path='mobile')
    def mobile_upload(self, request):
        '''
        Similar to batch create, but the mobile view is less punishing and allows for partial 
        successes. Will expect the information to be uploaded with both a server_id and a local device ID. 
        '''
        if request.user.role == 'client':
                raise PermissionDenied('You do not have permission to perform this action.')
        data = request.data
        if not isinstance(data, list):
            return Response({"detail": "Expected a list of interactions."}, status=400)
        ids_map = [] #map that compares server IDs to local IDs so the app knows what information was recorded
        errors = []
        for item in data:
            try:
                server_id = item.get('server_id')
                existing = None
                if server_id:
                    existing = Interaction.objects.filter(id=server_id).first()
                if existing: #auto update existing respondents (matched by ID no)
                    interaction_serializer = InteractionSerializer(
                        existing, data=item, context={'request': request}, partial=True
                    )
                else: #otherwise create new
                    interaction_serializer = InteractionSerializer(
                        data=item, context={'request': request}
                    )
                interaction_serializer.is_valid(raise_exception=True)
                interaction = interaction_serializer.save()
                ids_map.append({'local_id': item.get('local_id'), 'server_id': interaction.id })
            #catch and return any errors
            except Exception as err:
                errors.append({
                    'local_id': item.get('local_id'),
                    'message': str(err),
                })  
        return Response({
            "mappings": ids_map,
            "errors": errors
        }, status=status.HTTP_200_OK)


    ###=== FILE UPLOAD VIEWS ===###
    @action(detail=False, methods=['post'], url_path='template')
    def get_template(self, request):
        '''
        Action to generate an excel template that a user can input data into and upload for bulk
        uploading.
        '''
        #Check perms to make sure the user should have access to tempaltes
        user=request.user
        if not user.role in ['meofficer', 'manager', 'admin']:
            raise PermissionDenied('You do not have permission to access templates.')
        org_id = self.request.data.get('organization_id')
        task_ids = self.request.data.get('task_ids')
        if not task_ids or not org_id:
            raise serializers.ValidationError('Template requires a project and organization.')
            


        
        #pull user-friendly labels that users can view
        district_labels = [choice.label for choice in Respondent.District]
        sex_labels = [choice.label for choice in Respondent.Sex]
        age_range_labels = [choice.label for choice in Respondent.AgeRanges]
        kp_type_labels = [choice.label for choice in KeyPopulation.KeyPopulations]
        dis_labels = [dis.label for dis in DisabilityType.DisabilityTypes]
        auto_attr = [RespondentAttributeType.Attributes.PLWHIV, RespondentAttributeType.Attributes.KP, RespondentAttributeType.Attributes.PWD]
        special_attribute_labels = [attr.label for attr in RespondentAttributeType.Attributes if attr not in auto_attr]


        headers = [] #tracks what should go in the header

        #include all respondent information, except for a few meta_fields
        for field in Respondent._meta.get_fields():
            if field.auto_created:
                continue
            if field.name in ['uuid', 'created_by', 'created_at', 'updated_at', 'updated_by', 'flags', 'dummy_dob']:
                continue
            if hasattr(field, 'verbose_name'):
                verbose = field.verbose_name
            else:
                verbose = field.name

            field_info = {'header': verbose or field.name}
            if field.many_to_many:
                field_info['multiple'] = True
            else:
                field_info['multiple'] = False

            if field.name == 'is_anonymous':
                field_info['options'] = ['TRUE', 'FALSE']
            elif field.name == 'district':
                field_info['options'] = district_labels
            elif field.name == 'sex':
                field_info['options'] = sex_labels
            elif field.name == 'age_range':
                field_info['options'] = age_range_labels
            elif field.name == 'kp_status':
                field_info['options'] = kp_type_labels
            elif field.name == 'disability_status':
                field_info['options'] = dis_labels
            elif field.name == 'special_attribute':
                field_info['options'] = special_attribute_labels
            else:
                field_info['options'] = []
            headers.append(field_info)
        
        #push related columns for HIV status and pregnancy (these do not live in the respondent model, so we need to manually create these)
        headers.append({'header': 'HIV Status', 'options': ['HIV Positive', 'HIV Negative'], 'multiple': False})
        headers.append({'header': 'Date Positive', 'options': [], 'multiple': False})
        headers.append({'header': 'Pregnancy Began (Date)', 'options': [], 'multiple': False})
        headers.append({'header': 'Pregnancy Ended (Date)', 'options': [], 'multiple': False})

        #add cols for date of interaction and location
        headers.append({'header': 'Date of Interaction', 'options': [], 'multiple': False})
        headers.append({'header': 'Interaction Location', 'options': [], 'multiple': False})
        
    

        #create a header for each task
        for task_id in task_ids:
            task = Task.objects.filter(id=task_id).first()
            if not task.organization_id == org_id:
                raise ValidationError('This task does not belong to this organization.')
            if not task.assessment:
                raise ValidationError('You can only generate this template for assessments.')
            if user.role != 'admin':
                if task.organization != user.organization:
                    if not ProjectOrganization.objects.filter(parent_organization=user.organization, project_id=task.project_id, organization_id=org_id).exists(): #check if its a valid child for the project
                        raise PermissionDenied('You do not have permission to access this template.')
            
            for indicator in Indicator.objects.filter(assessment=task.assessment).order_by('order').all():
                if indicator.type == Indicator.Type.MULTI:
                    options = Option.objects.filter(indicator=indicator) if not indicator.match_options else Option.objects.filter(indicator=indicator.match_options)
                    for option in options.all():
                        header = f'{indicator.name}: {option.name} (Select All That Apply)'
                        categories = ['Yes', 'No']
                        headers.append({'header': header, 'options': categories})
                    continue
                header = indicator.name
                categories = []
                if indicator.type == Indicator.Type.INT:
                    header = header + ' (Enter a Number)'
                elif indicator.type == Indicator.Type.SINGLE:
                    header = header + ' (Select One)'
                    categories = [o.name for o in Option.objects.filter(indicator=indicator).all()]
                elif indicator.type == Indicator.Type.BOOL:
                    categories = ['Yes', 'No']
                headers.append({'header': header, 'options': categories})
        headers.append({'header': 'Comments', 'options': []})

        template_path = os.path.join(settings.BASE_DIR, 'respondents', 'static', 'respondents', 'upload-template.xlsx')
        wb = load_workbook(template_path)
        ws = wb.create_sheet('Data')
        options_sheet = wb.create_sheet('DropdownOptions')
        cols = list(excel_columns())

        #create the dropdown options sheet and hide it by default
        for c, header in enumerate(headers, 1):
            header_text = header['header']
            ws.cell(row=1, column=c, value=header_text)
            if len(header['options']) > 0:
                col = cols[c-1]
                for r, option in enumerate(header['options'], 1):
                    options_sheet[f'{col}{r}'] = str(option)
                dv = DataValidation(type="list", formula1=f"=DropdownOptions!${col}$1:${col}${len(header['options'])}", allow_blank=True)
                dv.add(f'{cols[c-1]}2:{cols[c-1]}1000')
                ws.add_data_validation(dv)
            else:
                continue
        options_sheet.sheet_state = 'hidden'

        #set metadata so the user doesn't have to specify the project/org again
        metadata_sheet = wb.create_sheet("Metadata")
        metadata_sheet["A1"] = "organization_id"
        metadata_sheet["A2"] = org_id
        metadata_sheet["B1"] = "task_ids"
        for i, task_id in enumerate(task_ids):
            metadata_sheet[f'B{i+2}'] = task_id
        metadata_sheet["C1"] = "number of tasks"
        metadata_sheet["C2"] = len(task_ids)

        metadata_sheet.sheet_state = 'hidden'
        metadata_sheet.protection.sheet = True
        metadata_sheet.protection.password = 'xQvzLit1@SS69'

        output = BytesIO()
        wb.save(output)
        output.seek(0)
        response = HttpResponse(output.read(), content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        filename = f'template_{date.today().strftime("%Y-%m-%d")}.xlsx'
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        return response

    @action(detail=False, methods=['POST'], url_path='upload')
    def post_template(self, request):
        '''
        Method for uploading the afforementioned template and converting it the data the system can use.
        '''

        user = request.user
        #custom errors/warnings
        errors = []
        warnings = []
        
        #check user has adequate role
        if user.role not in ['admin', 'meofficer', 'manager']:
            raise PermissionDenied('You do not have permission to access templates.')

        #pull the file and throw an error if its the wrong file type
        uploaded_file = request.FILES.get('file')
        if not uploaded_file:
            return Response({"detail": "No file was uploaded."}, status=status.HTTP_400_BAD_REQUEST)
        if not uploaded_file.name.endswith('.xlsx'):
            return Response({"detail": "Uploaded file must be an .xlsx Excel file."}, status=status.HTTP_400_BAD_REQUEST)
        
        #read metadata sheet that has project/organization info and throw an error if its missing or wrong
        try:
            wb = load_workbook(filename=uploaded_file)
            ws = wb['Metadata']
        except Exception:
            raise ValidationError("Unable to read 'Metadata' sheet. Please check the template.")

        try:
            org_id = int(ws['A2'].value)
        except (TypeError, ValueError):
            raise ValidationError("Organization ID must be numeric.")
        if not org_id:
            raise ValidationError("Template requires a valid organization ID.")
        
        if user.role != 'admin':
            #non admins should only have access to their org and child orgs
            if str(org_id) != str(user.organization_id): #if not their org then...
                if not ProjectOrganization.objects.filter(parent_organization=user.organization, organization_id=org_id).exists(): #check if its a valid child for the project
                    raise PermissionDenied('You do not have permission to access this template.')

        ws = wb['Data'] 
        headers = {}
        for row in ws.iter_rows(min_row=1, max_row=1):
            for cell in row:
                if cell.value:
                    header_name = str(cell.value).strip()
                    headers[header_name] = {
                        'column': cell.column,
                        'options': [],
                        'multiple': False
                    }

        #pull the labels used for the dropdown options (the same ones used to create the tempalte)
        #for simplicity/verification, we're treating everything as lowercase no spaces
        district_labels = [choice.label.lower().replace(' ', '') for choice in Respondent.District]
        sex_labels = [choice.label.lower().replace(' ', '') for choice in Respondent.Sex]
        age_range_labels = [choice.label.lower().replace(' ', '')  for choice in Respondent.AgeRanges]
        kp_type_labels = [choice.label.lower().replace(' ', '')  for choice in KeyPopulation.KeyPopulations]
        dis_labels = [dis.label.lower().replace(' ', '')  for dis in DisabilityType.DisabilityTypes]
        auto_attr = [RespondentAttributeType.Attributes.PLWHIV, RespondentAttributeType.Attributes.KP, RespondentAttributeType.Attributes.PWD]
        special_attribute_labels = [attr.label.lower().replace(' ', '') for attr in RespondentAttributeType.Attributes if attr not in auto_attr]
        def get_verbose(field_name):
            return Respondent._meta.get_field(field_name).verbose_name
        
        def expect_column(field_name, options=None, multiple=False):
            verbose = get_verbose(field_name)
            if verbose in headers:
                if options:
                    headers[verbose]['options'] = options
                headers[verbose]['multiple'] = multiple
            else:
                errors.append(f"Template is missing {verbose} column.")
        
        # make sure that all the required columns are present. If its missing one, the template was tampered with
        # and is not valid

        expect_column('id_no')
        expect_column('first_name')
        expect_column('last_name')
        expect_column('age_range', options=age_range_labels)
        expect_column('dob')
        expect_column('sex', options = sex_labels)
        expect_column('ward')
        expect_column('village')
        expect_column('district', options=district_labels)
        expect_column('citizenship')
        expect_column('email')
        expect_column('phone_number')
        expect_column('kp_status', options=kp_type_labels, multiple=True)
        expect_column('disability_status', options=dis_labels, multiple=True) 
        expect_column('special_attribute', options=special_attribute_labels, multiple=True)

        if not 'Date of Interaction' in headers:
            errors.append('Template is missing Date of Interaction column.')
        if not 'Interaction Location' in headers:
            errors.append('Template is missing Interaction Location column.')
        if not 'HIV Status' in headers:
            errors.append('Template is missing HIV Status column.')
        if not 'Date Positive' in headers:
            errors.append('Template is missing Date Positive column.')
        if not 'Pregnancy Began (Date)' in headers:
            errors.append('Template is missing Pregnant Began column.')
        if not 'Pregnancy Ended (Date)' in headers:
            errors.append('Template is missing Pregnant Ended column.')
        
        metadata_ws = wb['Metadata']
        data_ws = wb['Data']

        # Parse number of tasks
        try:
            num_tasks = int(metadata_ws['C2'].value)
        except (TypeError, ValueError):
            raise ValidationError("Task lengths must be numeric.")

        # Validate tasks
        task_ids = []
        assessments = set()

        def get_indicator_column(indicator, option=None):
            header = indicator.name
            if indicator.type == Indicator.Type.MULTI:
                header = f'{indicator.name}: {option} (Select All That Apply)'
            elif indicator.type == Indicator.Type.INT:
                header = header + ' (Enter a Number)'
            elif indicator.type == Indicator.Type.SINGLE:
                header = header + ' (Select One)'
            header = headers.get(header)
            if header:
                return header['column']
            return None
        
        for i in range(num_tasks):
            cell = metadata_ws[f'B{i+2}']  # B3, B4, B5...
            try:
                task_id = int(cell.value)
            except (TypeError, ValueError):
                errors.append(f"Task ID in row {cell.row} must be numeric.")
                continue
            task = Task.objects.filter(id=task_id).select_related('organization').first()
            
            if not task:
                errors.append(f"Task with ID {task_id} not found in metadata.")
                continue

            if str(task.organization_id) != str(org_id):
                errors.append("This task does not belong to this organization.")
                continue
            if not task.assessment:
                errors.append("This template is only valid for assessments.")
                continue
            # Role-based permissions
            if user.role != 'admin':
                if task.organization != user.organization and not ProjectOrganization.objects.filter(
                    parent_organization=user.organization,
                    project_id=task.project_id,
                    organization_id=org_id
                ).exists():
                    raise PermissionDenied("You do not have permission to use this template.")

            task_ids.append(task_id)
            assessments.add(task.assessment_id)
        # 2️⃣ Validate Indicators for Each Assessment
        for assessment_id in assessments:
            for indicator in Indicator.objects.filter(assessment_id=assessment_id).all():
                if indicator.type == Indicator.Type.MULTI:
                    options = Option.objects.filter(indicator=indicator.match_options) if indicator.match_options else Option.objects.filter(indicator=indicator)
                    for option in options.all():
                        col = get_indicator_column(indicator, option)
                        if not col:
                            errors.append(f'Missing column "{indicator.name}: {option.name} (Select All That Apply)"')
                else:
                    suffix = (
                        " (Select One)" if indicator.type == Indicator.Type.SINGLE else
                        " (Enter a Number)" if indicator.type == Indicator.Type.INT else
                        ""
                    )
                    col = get_indicator_column(indicator)
                    if not col:
                        errors.append(f'Missing column "{indicator.name}{suffix}"')


        
        #few helper functions to get us through the next step
        def get_cell_value(row, field_name):
            header = headers.get(get_verbose(field_name))
            if header:
                return row[header['column'] - 1]
            return None

        def get_column(field_name):
            header = headers.get(get_verbose(field_name))
            if header:
                return header['column']
            return None

        def get_indicator_value(row, indicator, option=None):
            header = indicator.name
            if indicator.type == Indicator.Type.MULTI:
                header = f'{indicator.name}: {option} (Select All That Apply)'
            elif indicator.type == Indicator.Type.INT:
                header = header + ' (Enter a Number)'
            elif indicator.type == Indicator.Type.SINGLE:
                header = header + ' (Select One)'
            header = headers.get(header)
            if header:
                return row[header['column'] - 1]
            return None

        def get_options(field_name):
            header = headers.get(get_verbose(field_name))
            if header:
                return header['options']
            return []
        
        if len(errors) > 0:
            return Response({'errors': errors, 'warnings': warnings,  }, status=status.HTTP_400_BAD_REQUEST)
        
        #track both new and existing respondents as we're going through each row
        created = []
        existing = []

        
        #loop through each row and collect/validate the data
        for i, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
            respondent = None

            #set unique array for tracking row errors/warnings
            row_errors = []
            row_warnings = []

            anon = is_truthy(get_cell_value(row, 'is_anonymous'))

            #check if this respondent already exists (mathcing id_no)
            id_no = get_cell_value(row, 'id_no') or None
            if id_no and not anon:
                respondent = Respondent.objects.filter(id_no=id_no).first()

            #make sure first_name/last_name are present if respondent is not anonymous
            first_name = get_cell_value(row, 'first_name') or None
            last_name = get_cell_value(row, 'last_name') or None

            if not anon and not first_name:
                row_errors.append(f"Respondent at column: {get_column('first_name')}, row: {i} requires a first name.")
            if not anon and not last_name:
                row_errors.append(f"Respondent at column: {get_column('last_name')}, row: {i} requires a last name.")

            #check that dob is present for non-anons and is a valid date
            dob = get_cell_value(row, 'dob') or None
            if anon:
                dob = None
            else:
                if dob:
                    parsed = valid_excel_date(dob)
                    if not parsed:
                        row_errors.append(f"Date of birth {dob} at column: {get_column('dob')}, row: {i} is invalid. Double check the format and make sure that it is not in the future")
                    dob = parsed
                else:
                    row_errors.append(f"Date of birth at column: {get_column('dob')}, row: {i} is required for non-anonymous respondents.")
            
            #check that sex is  a valid value
            sex = get_cell_value(row, 'sex')
            if sex:
                sex=sex.lower().replace(' ', '')
                if not sex in get_options('sex'):
                    row_errors.append(f"Sex at column: {get_column('sex')}, row: {i} is not a valid choice.")
            else:
                row_errors.append(f"Sex at column: {get_column('sex')}, row: {i} is required.")
            
            #check that age range is present for anons (ignore for non-anons since we alread have the DOB)
            age_range = get_cell_value(row, 'age_range')
            if anon:
                if age_range:
                    age_range = age_range.lower().replace(' ', '')
                    if not age_range in get_options('age_range'):
                        row_errors.append(f"Age Range value at column: {get_column('age_range')}, row: {i} is not a valid choice.")
                else:
                    row_errors.append(f"Age range at column: {get_column('sex')}, row: {i} is required for anonymous respondents.")
            
            #get ward, optional
            ward = get_cell_value(row, 'ward') or None
            
            #get village and district (required, district must match a value)
            village = get_cell_value(row, 'village') or None
            if not village:
                row_errors.append(f"Village at column: {get_column('village')}, row: {i} is required for all respondents.")

            district = get_cell_value(row, 'district') or None
            if district:
                district = district.lower().replace(' ', '')
                if not district in get_options('district'):
                    row_errors.append(f"District at column: {get_column('district')}, row: {i} is not a valid choice.")
            else:
                row_errors.append(f"District at column: {get_column('district')}, row: {i} is required.")
            
            #get the citizenship, if left blank, assume they are a citizen but throw a warning to make sure
            citizenship = get_cell_value(row, 'citizenship') or None
            if not citizenship:
                citizenship = 'BW'
                row_warnings.append(f"Citizenship at column: {get_column('citizenship')}, row: {i} is required for all respondents. This value will default to BW. If this is incorrect, please check this field again.")
            #if not blank, verify a proper country name was provided
            else:
                try:
                    citizenship = pycountry.countries.lookup(citizenship).alpha_2
                except LookupError:
                    row_errors.append(f"Citizenship {citizenship} at column: {get_column('citizenship')}, row: {i} is not a valid country name/code.")
            
            #get/validate email and phone if provided
            email = get_cell_value(row, 'email') or None
            if email and not is_email(email):
                row_warnings.append(f"Email at column: {get_column('email')}, row: {i} is not a valid choice.")
                email = None
            phone_number = get_cell_value(row, 'phone_number') or None
            if phone_number and not is_phone_number(phone_number):
                row_warnings.append(f"Phone Number at column: {get_column('phone_number')}, row: {i} is not a valid choice.")
                phone_number = None


            def get_choice_key_from_label(choices, label):
                if not label:
                    return None
                for key, value in choices:
                    if value.lower().replace(' ', '') == label.lower().replace(' ', ''):
                        return key
                return None
            
            sex = get_choice_key_from_label(Respondent.Sex.choices, sex)
            district = get_choice_key_from_label(Respondent.District.choices, district)
            if age_range and not dob:
                age_range = get_choice_key_from_label(Respondent.AgeRanges.choices, age_range)
            elif dob:
                age_range = None
            #validate our m2m fields
            kp_types = []
            kp_status_names_raw = get_cell_value(row, 'kp_status') or ''
            if kp_status_names_raw:
                # Clean and split
                cleaned = kp_status_names_raw.replace(' ', '').lower()
                input_kp_names = set(re.split(r'[,:;]', cleaned))

                valid_labels = get_options('kp_status')
                valid_lookup = {label.replace(' ', '').lower(): label for label in valid_labels}

                matched = [name for name in input_kp_names if name in valid_lookup]
                invalid_kp = input_kp_names - set(matched)

                if invalid_kp:
                    row_warnings.append(
                        f"Invalid key population statuses at row {i}: {', '.join(invalid_kp)}"
                    )
                for cleaned_name in matched:
                    key = get_choice_key_from_label(KeyPopulation.KeyPopulations.choices, cleaned_name)
                    if not key:
                        continue
                    kp, _ = KeyPopulation.objects.get_or_create(name=key)
                    kp_types.append(kp)

            disability_types = []
            disability_status_names_raw = get_cell_value(row, 'disability_status') or ''
            if disability_status_names_raw:
                # Clean and split
                cleaned = disability_status_names_raw.replace(' ', '').lower()
                input_disability_names = set(re.split(r'[,:;]', cleaned))

                valid_labels = get_options('disability_status')
                valid_lookup = {label.replace(' ', '').lower(): label for label in valid_labels}

                matched = [name for name in input_disability_names if name in valid_lookup]
                invalid_disability = input_disability_names - set(matched)

                if invalid_disability:
                    row_warnings.append(
                        f"Invalid disability statuses at row {i}: {', '.join(invalid_disability)}"
                    )
                
                for cleaned_name in matched:
                    key = get_choice_key_from_label(DisabilityType.DisabilityTypes.choices, cleaned_name)
                    if not key:
                        continue
                    dis, _ = DisabilityType.objects.get_or_create(name=key)
                    disability_types.append(dis)

            special_attr_names_raw = get_cell_value(row, 'special_attribute') or ''
            attr_types = []
            if special_attr_names_raw:
                # Clean and split
                cleaned = special_attr_names_raw.replace(' ', '').lower()
                input_attr_names = set(re.split(r'[,:;]', cleaned))

                valid_labels = get_options('special_attribute')
                valid_lookup = {label.replace(' ', '').lower(): label for label in valid_labels}

                matched = [name for name in input_attr_names if name in valid_lookup]
                invalid_attr = input_attr_names - set(matched)

                if invalid_attr:
                    row_warnings.append(
                        f"Invalid respondent attribute at row {i}: {', '.join(invalid_attr)}"
                    )
                
                for cleaned_name in matched:
                    key = get_choice_key_from_label(RespondentAttributeType.Attributes.choices, cleaned_name)
                    if not key:
                        continue
                    attr, _ = RespondentAttributeType.objects.get_or_create(name=key)
                    attr_types.append(attr)

            #get/validate HIV Status cols
            hs_col = headers['HIV Status']['column']-1 
            hiv_status = row[hs_col] if len(row) > hs_col else None
            dp_col = headers['Date Positive']['column']-1 
            date_positive = row[dp_col] if len(row) > dp_col else None
            if hiv_status:
                if hiv_status.lower().replace(' ', '') == 'yes':
                    hiv_status = True
                    if not date_positive:
                        row_warnings.append(f"HIV Status at row {i} does not have a date positive. We will automatically set the date as today. Please double check this entry.")
                        date_positive = date.today()
                    parsed = valid_excel_date(date_positive)
                    if not parsed:
                        row_errors.append(
                            f"Date positive '{date_positive}' at column: {dp_col}, row: {i} is invalid."
                            "Double check the format and make sure that it is not in the future."
                        )
                    else:
                        date_positive = parsed
                else:
                    hiv_status=None
                    date_positive = None
            else:
                hiv_status = None
                date_positive = None
            
            #get/validate pregnancy term_began/term_ended cols
            tb_col = headers['Pregnancy Began (Date)']['column']-1 
            term_began = row[tb_col] if len(row) > tb_col else None
            te_col = headers['Pregnancy Ended (Date)']['column']-1 
            term_ended = row[te_col] if len(row) > te_col else None
            if not term_ended and not term_began:
                term_ended = None
                term_began = None
            else:
                if term_ended and not term_began:
                    row_errors.append(
                            f"Pregnancy at row {i} requires a start date."
                        )
                else:
                    if term_began:
                        parsed = valid_excel_date(term_began)
                        if not parsed:
                            row_errors.append(
                                f"Pregnancy Term Began '{term_began}' at column: {tb_col}, row: {i} is invalid."
                                "Double check the format and make sure that it is not in the future."
                            )
                        else:
                            term_began = parsed
                        if term_ended:
                            parsed = valid_excel_date(term_ended)
                            if not parsed:
                                row_errors.append(
                                    f"Pregnancy Term Began '{term_began}' at column: {tb_col}, row: {i} is invalid."
                                    "Double check the format and make sure that it is not in the future."
                                )
                            else:
                                term_ended = parsed
                        else:
                            term_ended = None
            #if there are any errors up to this point, the user needs to verify the respondent before any any data is recorded
            if len(row_errors) > 0:
                row_errors.append("This respondent and their interactions will not be saved until these errors are fixed")
                errors.extend(row_errors)
                warnings.extend(row_warnings)
                continue
            
            #mock a request so the Respondent serializer can take us the rest of the way
            class FakeRequest:
                def __init__(self, user):
                    self.user = user

            def process_row(row_data, request_user):
                serializer = RespondentSerializer(data=row_data, context={"request": FakeRequest(request_user)})
                if serializer.is_valid():
                    respondent = serializer.save()
                    return respondent, None
                else:
                    return None, serializer.errors
            #append the created data to our master list if new 
            if not respondent:
                respondent_data = upload = {
                    'id_no': id_no,
                    'is_anonymous': anon,
                    'first_name': first_name,
                    'last_name': last_name,
                    'ward': ward,
                    'village': village,
                    'district': district,
                    'sex': sex,
                    'dob': dob,
                    'age_range' : age_range,
                    'citizenship': citizenship,
                    'email': email,
                    'phone_number': phone_number,
                    'kp_status_names': [kp.name for kp in kp_types],
                    'disability_status_names': [d.name for d in disability_types],
                    'special_attribute_names': [attr.name for attr in attr_types],
                    'hiv_status_data': {'hiv_positive': hiv_status, 'date_positive': date_positive},
                    'pregnancy_data': [{'term_began': term_began, 'term_ended': term_ended}],
                }
                respondent, err = process_row(respondent_data, user)
                if err:
                    row_errors.append({"row": i + 1, "errors": err})
                else:
                    created.append(respondent)
            #if existing append them to the existing list, we'll use this later in the frontend to compare users
            else:
                ex_stat = HIVStatus.objects.filter(respondent=respondent).first()
                upload = {
                    'is_anonymous': anon,
                    'first_name': first_name,
                    'last_name': last_name,
                    'ward': ward,
                    'village': village,
                    'district': district,
                    'sex': sex,
                    'dob': dob,
                    'age_range' : age_range,
                    'citizenship': citizenship,
                    'email': email,
                    'phone_number': phone_number,
                    'kp_status_names': sorted([kp.name for kp in kp_types]),
                    'disability_status_names': sorted([d.name for d in disability_types]),
                    'special_attribute_names': sorted([attr.name for attr in attr_types]),
                    'hiv_status_data': {'hiv_positive': hiv_status, 'date_positive': date_positive},
                    'pregnancy_data': [{'term_began': term_began, 'term_ended': term_ended}],
                }
                auto_attr = [RespondentAttributeType.Attributes.PLWHIV, RespondentAttributeType.Attributes.KP, RespondentAttributeType.Attributes.PWD]
                in_db = {
                    'is_anonymous': respondent.is_anonymous,
                    'first_name': respondent.first_name,
                    'last_name': respondent.last_name,
                    'ward': respondent.ward,
                    'village': respondent.village,
                    'district': respondent.district,
                    'sex': respondent.sex,
                    'dob': respondent.dob,
                    'age_range' : respondent.age_range,
                    'citizenship': respondent.citizenship,
                    'email': respondent.email,
                    'phone_number': respondent.phone_number,
                    'kp_status_names': sorted([kp.name for kp in respondent.kp_status.all()]),
                    'disability_status_names': sorted([d.name for d in respondent.disability_status.all()]),
                    'special_attribute_names': sorted([
                        attr.name for attr in respondent.special_attribute.all()
                        if getattr(attr, 'name', None) not in auto_attr   # adjust field name if needed
                    ]),
                    'hiv_status_data': {'hiv_positive': ex_stat.hiv_positive if ex_stat else None, 'date_positive': ex_stat.date_positive if ex_stat else None},
                }
                # Remove pregnancy before comparing (optional)
                upload_preg = upload.pop('pregnancy_data', None)
                if not anon:
                    in_db['age_range'] = None
                if upload != in_db:
                    existing.append({'id': respondent.id, 'upload': upload, 'in_database': in_db})

            #get date of interaction and make sure its legit
            doi_col = headers['Date of Interaction']['column']-1 
            interaction_date = row[doi_col] if len(row) > doi_col else None

            if interaction_date:
                parsed = valid_excel_date(interaction_date)
                if not parsed:
                    row_errors.append(
                        f"Date of interaction '{interaction_date}' at column: {doi_col}, row: {i} is invalid. "
                        "Double check the format and make sure that it is not in the future."
                    )
                interaction_date = parsed
            else:
                row_errors.append(
                    f"Date of interaction at column: {doi_col}, row: {i} is required. "
                )
            loc_col = headers['Interaction Location']['column']-1 
            comments = get_cell_value(row, 'comments') or None
            #make sure a location is provided
            interaction_location = row[loc_col] if len(row) > loc_col else None
            if not interaction_location:
                row_errors.append(
                    f"Date of location at column: {doi_col}, row: {i} is required. "
                )
            
            #any errors with date/location, don't record any interactions until its fixed
            if len(row_errors) > 0:
                row_errors.append(f"This respondent has been saved, but none of their interactions will until this date is fixed.")
                errors.extend(row_errors)
                warnings.extend(row_warnings)
                continue
            
            #otherwise, loop throught the task columns
            
            indicator_columns = {}
            for assessment_id in assessments:
                response_data = {}
                for indicator in Indicator.objects.filter(assessment_id=assessment_id).order_by('order'):
                    print(indicator.name)
                    col = None
                    val = None
                    if indicator.type == Indicator.Type.MULTI:
                        val = []
                        for option in Option.objects.filter(indicator=indicator) if not indicator.match_options else Option.objects.filter(indicator=indicator.match_options):
                            col = get_indicator_column(indicator, option)
                            o_val = str(get_indicator_value(row, indicator, option))
                            o_val = o_val.lower().replace(' ', '')
                            if o_val in ['', 'no', 'none', 'na', 'n/a', 'false', 'unsure', 'maybe']:
                                continue
                            val.append(option.id)
                        if len(val) == 0 and indicator.allow_none:
                            val = ['none']
                    else:
                        col = get_indicator_column(indicator)
                        val = str(get_indicator_value(row, indicator))
                        val = val.lower().replace(' ', '') if indicator.type != Indicator.Type.TEXT else val
                        if val == 'none' and indicator.type == Indicator.Type.SINGLE and indicator.allow_none:
                            val == 'none'
                        elif val in ['', 'none', 'na', 'n/a', 'unsure', 'maybe']:
                            continue
                    if indicator.type == Indicator.Type.SINGLE:
                        valid_map = {
                            o.name.lower().replace(' ', ''): o.id
                            for o in Option.objects.filter(indicator=indicator).all()
                        }
                        if val not in valid_map:
                            row_warnings.append(
                                f'Could not parse value at column: {col}, row: {i}. Please enter a valid option.'
                            )
                        else:
                            val = valid_map[val]

                    if indicator.type == Indicator.Type.BOOL:
                        if val in ['yes', 'true', '1']:
                            val = True
                        elif val in ['no', 'false', '0']:
                            val = False
                        else:
                            row_warnings.append(f'Could not parse value at column: {col}, row: {i}. Please enter "yes" or "no".')
                    if indicator.type == Indicator.Type.INT:
                        try:
                            numeric_component = int(val)
                            if numeric_component < 0:
                                row_warnings.append(f'Number at column: {col}, row: {i} must be greater than 0.')
                                continue
                        except (ValueError, TypeError):
                            row_warnings.append(f'Number at column: {col}, row: {i} is not a valid number.')
                            continue
                    indicator_columns[str(indicator.id)] = col
                    response_data[str(indicator.id)] = { 'value': val }
                
                print(response_data)
                # run once per assessment
                lookup_fields = {
                    'respondent': respondent,
                    'interaction_date': interaction_date,
                    'task': task,
                }

                # Try to fetch the existing interaction
                instance = Interaction.objects.filter(**lookup_fields).first()

                # Pass instance to serializer if it exists (update), otherwise it will create, let it take us the rest of the way
                serializer = self.get_serializer(
                    instance=instance,
                    data={
                        'respondent_id': respondent.id,
                        'interaction_date': interaction_date,
                        'interaction_location': interaction_location,
                        'task_id': task.id,
                        'response_data': response_data,
                        'comments': comments,
                    },
                    context={'request': request, 'respondent': respondent}
                )
                try:
                    serializer.is_valid(raise_exception=True)
                    serializer.save()
                except ValidationError as e:
                    # Flatten error details for easier reading
                    error_details = serializer.errors
                    details = error_details.get("details", {})
                    indicator_id = str(details.get("indicator_id", "?"))  # convert to str
                    col = indicator_columns.get(indicator_id, "?")

                    for field, msgs in error_details.items():
                        if field == "details":
                            continue  # skip metadata key
                        if isinstance(msgs, list):
                            for msg in msgs:
                                row_errors.append(f"Row {i}, Column {col}, Field '{field}': {str(msg)}")
                        else:
                            row_errors.append(f"Row {i}, Column {col}, Field '{field}': {str(msgs)}")
            
            #push our row errors to the main append
            errors.extend(row_errors)
            warnings.extend(row_warnings)
        print('warnings', warnings)
        print('errors', errors)
        #wow, you made it. Congrats!
        return Response({'errors': errors, 'warnings': warnings, 'created': len(created), 'conflicts': existing }, status=status.HTTP_200_OK)